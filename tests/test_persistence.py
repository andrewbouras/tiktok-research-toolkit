import json
import shutil
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

import tiktok_research.dashboard.app as dashboard_app
from tiktok_research.workspace import init_demo_workspace


def build_complete_payload(misinfo_score: str = "3") -> dict:
    payload = {
        "Pre_Watch_Prediction": "1",
        "Thumbnail_White_Coat": "yes",
        "Thumbnail_Clinical_Setting": "no",
        "Thumbnail_Text_Overlay": "1",
        "Thumbnail_Creator_Visible": "1",
        "B1_Creator_Type": "healthcare_professional",
        "B2_Definition": "1",
        "B2_Causes": "1",
        "B2_PelvicFloor": "1",
        "B2_Behavioral": "0",
        "B2_Medical": "0",
        "B2_Surgical": "0",
        "B2_Products": "0",
        "B2_SeeDr": "1",
        "B2_Other": "0",
        "B3_Misinfo_Score": misinfo_score,
        "DQ1_Aims": "4",
        "DQ2_Achieves": "4",
        "DQ3_Relevant": "4",
        "DQ4_Sources": "2",
        "DQ5_Date": "2",
        "DQ6_Balanced": "3",
        "DQ7_AddlSources": "1",
        "DQ8_Uncertainty": "1",
        "DQ9_HowWorks": "4",
        "DQ10_Benefits": "3",
        "DQ11_Risks": "2",
        "DQ12_NoTreat": "1",
        "DQ13_QoL": "3",
        "DQ14_Options": "2",
        "DQ15_SharedDec": "4",
        "DQ16_Overall": "3",
    }
    if misinfo_score in {"4", "5"}:
        payload.update(
            {
                "B4_Misinfo_Details": "Incorrectly claims Kegels cure all incontinence.",
                "Misinfo_Types": ["TREAT-WRONG", "EXAG"],
                "Clinical_Domain": ["SUI", "pelvic_floor"],
                "B6_Notes": "Needs PI review for an overgeneralized treatment claim.",
                "Flag_For_Review": "1",
            }
        )
    return payload


class PersistenceTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.temp_dir = Path(tempfile.mkdtemp(prefix="research-dashboard-tests-"))
        init_demo_workspace(cls.temp_dir)
        cls.app = dashboard_app.create_app(cls.temp_dir)
        cls.client = cls.app.test_client()

    @classmethod
    def tearDownClass(cls) -> None:
        shutil.rmtree(cls.temp_dir, ignore_errors=True)

    def setUp(self) -> None:
        with self.app.app_context():
            conn = dashboard_app.get_db()
            conn.execute("DELETE FROM ratings")
            conn.commit()

    def load_responses(self, slug: str, video_id: str) -> dict:
        with self.app.app_context():
            conn = dashboard_app.get_db()
            researcher = dashboard_app.fetch_researcher_by_slug(conn, slug)
            row = conn.execute(
                """
                SELECT responses_json
                FROM ratings
                WHERE researcher_id = ? AND video_id = ?
                """,
                (researcher["id"], video_id),
            ).fetchone()
            self.assertIsNotNone(row)
            return json.loads(row["responses_json"])

    def test_researcher_route_starts_with_first_incomplete_video(self) -> None:
        response = self.client.get("/researchers/reviewer-a", follow_redirects=False)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/researchers/reviewer-a/videos/VID001")

    def test_researcher_route_resumes_most_recent_in_progress_video(self) -> None:
        partial = {
            "Pre_Watch_Prediction": "1",
            "Thumbnail_White_Coat": "yes",
            "Thumbnail_Clinical_Setting": "no",
            "Thumbnail_Text_Overlay": "1",
        }
        self.client.post(
            "/api/researchers/reviewer-a/videos/VID005",
            json={"responses": partial},
        )
        response = self.client.get("/researchers/reviewer-a", follow_redirects=False)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/researchers/reviewer-a/videos/VID005")

    def test_completed_video_advances_to_next_video_in_workbook_order(self) -> None:
        self.client.post(
            "/api/researchers/reviewer-a/videos/VID001",
            json={"responses": build_complete_payload()},
        )
        response = self.client.get("/researchers/reviewer-a", follow_redirects=False)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/researchers/reviewer-a/videos/VID002")

    def test_relevant_answers_persist_across_multiple_snapshot_saves(self) -> None:
        snapshot_one = {
            "Pre_Watch_Prediction": "1",
            "Thumbnail_White_Coat": "yes",
            "Thumbnail_Clinical_Setting": "no",
            "Thumbnail_Text_Overlay": "1",
            "Thumbnail_Creator_Visible": "1",
        }
        snapshot_two = {
            **snapshot_one,
            "B1_Creator_Type": "healthcare_professional",
            "B2_Definition": "1",
            "B2_Causes": "1",
            "B2_PelvicFloor": "1",
            "B2_Behavioral": "0",
            "B2_Medical": "0",
            "B2_Surgical": "0",
            "B2_Products": "0",
            "B2_SeeDr": "1",
            "B2_Other": "0",
        }
        self.client.post(
            "/api/researchers/reviewer-a/videos/VID004",
            json={"responses": snapshot_one},
        )
        self.client.post(
            "/api/researchers/reviewer-a/videos/VID004",
            json={"responses": snapshot_two},
        )
        responses = self.load_responses("reviewer-a", "VID004")
        self.assertEqual(responses["Pre_Watch_Prediction"], "1")
        self.assertEqual(responses["Thumbnail_White_Coat"], "yes")
        self.assertEqual(responses["B1_Creator_Type"], "healthcare_professional")
        self.assertEqual(responses["B2_SeeDr"], "1")

    def test_irrelevant_conditional_answers_are_cleared_on_resave(self) -> None:
        self.client.post(
            "/api/researchers/reviewer-a/videos/VID005",
            json={"responses": build_complete_payload("4")},
        )
        self.client.post(
            "/api/researchers/reviewer-a/videos/VID005",
            json={"responses": build_complete_payload("3")},
        )
        responses = self.load_responses("reviewer-a", "VID005")
        self.assertEqual(responses["B3_Misinfo_Score"], "3")
        self.assertNotIn("B4_Misinfo_Details", responses)
        self.assertNotIn("Misinfo_Types", responses)
        self.assertNotIn("Clinical_Domain", responses)
        self.assertNotIn("B6_Notes", responses)


    def test_backup_snapshot_is_written_after_save(self) -> None:
        self.client.post(
            "/api/researchers/reviewer-a/videos/VID003",
            json={"responses": build_complete_payload()},
        )
        backup_files = sorted((dashboard_app.BACKUP_DIR).glob("*.sqlite3"))
        self.assertTrue(backup_files)

    def test_custom_physician_can_be_deleted_without_touching_defaults(self) -> None:
        self.client.post(
            "/researchers",
            data={"display_name": "Test Physician"},
            follow_redirects=False,
        )
        with self.app.app_context():
            conn = dashboard_app.get_db()
            created = dashboard_app.fetch_researcher_by_slug(conn, "test-physician")
            self.assertIsNotNone(created)

        response = self.client.post(
            "/researchers/test-physician/delete",
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/")

        with self.app.app_context():
            conn = dashboard_app.get_db()
            deleted = dashboard_app.fetch_researcher_by_slug(conn, "test-physician")
            self.assertIsNone(deleted)
            self.assertIsNotNone(dashboard_app.fetch_researcher_by_slug(conn, "reviewer-a"))

    def test_default_physician_delete_is_blocked(self) -> None:
        response = self.client.post(
            "/researchers/reviewer-a/delete",
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 403)

    def test_field_order_matches_workbook_columns(self) -> None:
        workbook = load_workbook(dashboard_app.WORKBOOK_PATH, data_only=False)
        try:
            sheet = workbook["Reviewer_A"]
            workbook_fields = [
                sheet.cell(1, column).value
                for column in range(9, sheet.max_column + 1)
                if sheet.cell(1, column).value != "DISCERN_Total"
            ]
        finally:
            workbook.close()

        self.assertEqual(workbook_fields, dashboard_app.FIELD_ORDER)

    def test_score_4_or_5_followup_fields_match_workbook_logic(self) -> None:
        followup_fields = {
            "B4_Misinfo_Details",
            "Misinfo_Types",
            "Clinical_Domain",
            "B6_Notes",
        }
        for field_id in followup_fields:
            requirement = dashboard_app.FIELD_INDEX[field_id].get("required_when")
            self.assertIsNotNone(requirement)
            self.assertEqual(requirement["field"], "B3_Misinfo_Score")
            self.assertEqual(requirement["values"], ["4", "5"])

    def test_prewatch_unclear_option_is_available_and_documented(self) -> None:
        prewatch = dashboard_app.FIELD_INDEX["Pre_Watch_Prediction"]
        self.assertIn({"value": "2", "label": "Unclear from thumbnail alone"}, prewatch["options"])
        self.assertIn("2 = Unclear from thumbnail alone", prewatch["help_text"])
        self.assertEqual(prewatch["entry_hint"], "0 / 1 / 2")

    def test_discern_help_is_loaded_from_workbook_reference(self) -> None:
        dq1 = dashboard_app.FIELD_INDEX["DQ1_Aims"]
        dq16 = dashboard_app.FIELD_INDEX["DQ16_Overall"]
        self.assertIn("1=Not at all", dq1["help_text"])
        self.assertEqual(dq1["scale_hint"], "Workbook scale: 1 = Not at all, 5 = Fully.")
        self.assertIn("what the video is about", dq1["reference_note"])
        self.assertEqual(dq16["scale_hint"], "Workbook scale: 1 = Very poor, 5 = Excellent.")

    def test_misinfo_reference_rows_load_from_workbook(self) -> None:
        misinfo_section = next(section for section in dashboard_app.FIELD_SECTIONS if section["key"] == "misinfo")
        self.assertEqual(len(misinfo_section["score_reference"]), 5)
        self.assertEqual(misinfo_section["score_reference"][0]["score"], "1")
        self.assertIn("Directly contradicts guidelines", misinfo_section["score_reference"][-1]["definition"])


if __name__ == "__main__":
    unittest.main()
