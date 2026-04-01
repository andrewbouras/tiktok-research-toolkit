import json
import shutil
import sqlite3
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

import tiktok_research.sample_replacement as manage_sample


def build_workbook(path: Path, slot_url: str) -> None:
    workbook = Workbook()
    reviewer_a = workbook.active
    reviewer_a.title = "Reviewer_A"
    reviewer_b = workbook.create_sheet("Reviewer_B")

    for sheet in (reviewer_a, reviewer_b):
        headers = manage_sample.METADATA_COLUMNS + [
            "Pre_Watch_Prediction",
            "Thumbnail_White_Coat",
            "DISCERN_Total",
        ]
        for index, header in enumerate(headers, start=1):
            sheet.cell(1, index).value = header
        sheet.cell(4, 1).value = "VID001"
        sheet.cell(4, 2).value = "https://www.tiktok.com/@used/video/1111111111111111111"
        sheet.cell(5, 1).value = "VID002"
        sheet.cell(5, 2).value = slot_url
        sheet.cell(5, 3).value = 123
        sheet.cell(5, 4).value = 45
        sheet.cell(5, 5).value = 6
        sheet.cell(5, 6).value = 7
        sheet.cell(5, 7).value = "2025-03-31"
        sheet.cell(5, 8).value = "@oldcreator"
        sheet.cell(5, 9).value = "1"
        sheet.cell(5, 10).value = "yes"

    workbook.save(path)
    workbook.close()


def build_database(path: Path) -> None:
    conn = sqlite3.connect(path)
    conn.executescript(
        """
        CREATE TABLE videos (
            id TEXT PRIMARY KEY,
            tiktok_url TEXT NOT NULL,
            tiktok_numeric_id TEXT,
            views INTEGER,
            likes INTEGER,
            comments INTEGER,
            shares INTEGER,
            post_date TEXT,
            creator_handle TEXT,
            description TEXT,
            transcript TEXT,
            source_row INTEGER
        );
        CREATE TABLE researchers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            slug TEXT UNIQUE NOT NULL,
            display_name TEXT UNIQUE NOT NULL,
            created_at TEXT,
            updated_at TEXT
        );
        CREATE TABLE ratings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            researcher_id INTEGER NOT NULL,
            video_id TEXT NOT NULL,
            responses_json TEXT NOT NULL DEFAULT '{}',
            progress_pct REAL NOT NULL DEFAULT 0,
            is_complete INTEGER NOT NULL DEFAULT 0,
            flag_for_review INTEGER NOT NULL DEFAULT 0,
            started_at TEXT,
            updated_at TEXT,
            completed_at TEXT,
            UNIQUE(researcher_id, video_id)
        );
        """
    )
    conn.execute(
        """
        INSERT INTO videos (
            id, tiktok_url, tiktok_numeric_id, views, likes, comments, shares,
            post_date, creator_handle, description, transcript, source_row
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "VID002",
            "https://www.tiktok.com/@oldcreator/video/2222222222222222222",
            "2222222222222222222",
            123,
            45,
            6,
            7,
            "2025-03-31",
            "@oldcreator",
            "old description",
            "old transcript",
            5,
        ),
    )
    conn.execute(
        "INSERT INTO researchers (slug, display_name) VALUES (?, ?)",
        ("test", "Test"),
    )
    conn.execute(
        """
        INSERT INTO ratings (researcher_id, video_id, responses_json, progress_pct)
        VALUES (1, 'VID002', ?, 50.0)
        """,
        (json.dumps({"Pre_Watch_Prediction": "1"}),),
    )
    conn.commit()
    conn.close()


class SampleReplacementTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = Path(tempfile.mkdtemp(prefix="sample-replacement-tests-"))
        self.v2_workbook = self.temp_dir / "TikTok_UI_Video_Coding_Template_v2.xlsx"
        self.orig_workbook = self.temp_dir / "TikTok_UI_Video_Coding_Template.xlsx"
        build_workbook(self.v2_workbook, "https://www.tiktok.com/@oldcreator/video/2222222222222222222")
        build_workbook(self.orig_workbook, "https://www.tiktok.com/@oldcreator/video/2222222222222222222")

        self.database_path = self.temp_dir / "research_dashboard.sqlite3"
        build_database(self.database_path)

        self.collected_path = self.temp_dir / "collected_videos.json"
        self.collected_path.write_text(
            json.dumps(
                [
                    {
                        "id": 1111111111111111111,
                        "username": "used",
                        "view_count": 400,
                        "like_count": 30,
                        "comment_count": 3,
                        "share_count": 2,
                        "create_time": 1743379200,
                        "video_description": "already used",
                    },
                    {
                        "id": 3333333333333333333,
                        "username": "replacement",
                        "view_count": 999,
                        "like_count": 88,
                        "comment_count": 7,
                        "share_count": 5,
                        "create_time": 1743465600,
                        "video_description": "replacement description",
                    },
                ]
            )
        )

        self.transcripts_dir = self.temp_dir / "transcripts"
        self.transcripts_dir.mkdir()
        (self.transcripts_dir / "VID002.txt").write_text("old transcript")

        self.llm_scores_dir = self.temp_dir / "llm_scores" / "VID002"
        self.llm_scores_dir.mkdir(parents=True)
        (self.llm_scores_dir / "gpt.json").write_text("{}")

        self.config = manage_sample.ProjectConfig(
            workbook_paths=(self.v2_workbook, self.orig_workbook),
            database_path=self.database_path,
            candidate_pool_paths=(self.collected_path,),
            transcripts_dir=self.transcripts_dir,
            llm_scores_dir=self.temp_dir / "llm_scores",
            archive_root=self.temp_dir / "archives",
            replacement_log_path=self.temp_dir / "replacement_log.jsonl",
        )

    def tearDown(self) -> None:
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_list_reserve_candidates_excludes_used_ids(self) -> None:
        candidates = manage_sample.list_reserve_candidates(self.config)
        self.assertEqual(len(candidates), 1)
        self.assertEqual(candidates[0]["tiktok_numeric_id"], "3333333333333333333")

    def test_replace_slot_updates_workbooks_database_and_archives(self) -> None:
        result = manage_sample.replace_slot("VID002", "3333333333333333333", self.config)

        for workbook_path in (self.v2_workbook, self.orig_workbook):
            workbook = load_workbook(workbook_path, data_only=True)
            try:
                for sheet_name in ("Reviewer_A", "Reviewer_B"):
                    sheet = workbook[sheet_name]
                    self.assertEqual(sheet.cell(5, 2).value, "https://www.tiktok.com/@replacement/video/3333333333333333333")
                    self.assertEqual(sheet.cell(5, 3).value, 999)
                    self.assertEqual(sheet.cell(5, 8).value, "@replacement")
                    self.assertIsNone(sheet.cell(5, 9).value)
                    self.assertIsNone(sheet.cell(5, 10).value)
            finally:
                workbook.close()

        conn = sqlite3.connect(self.database_path)
        conn.row_factory = sqlite3.Row
        try:
            video_row = conn.execute("SELECT * FROM videos WHERE id = 'VID002'").fetchone()
            self.assertEqual(video_row["tiktok_numeric_id"], "3333333333333333333")
            self.assertEqual(video_row["creator_handle"], "@replacement")
            self.assertEqual(video_row["description"], "replacement description")
            self.assertEqual(video_row["transcript"], "")
            rating_count = conn.execute("SELECT COUNT(*) FROM ratings WHERE video_id = 'VID002'").fetchone()[0]
            self.assertEqual(rating_count, 0)
        finally:
            conn.close()

        self.assertFalse((self.transcripts_dir / "VID002.txt").exists())
        self.assertFalse((self.temp_dir / "llm_scores" / "VID002").exists())
        self.assertTrue(result["transcript_archive"])
        self.assertTrue(result["llm_scores_archive"])
        self.assertTrue(Path(result["database_backup"]).exists())
        self.assertTrue(Path(result["workbook_backups"][0]).exists())
        self.assertTrue(self.config.replacement_log_path.exists())
