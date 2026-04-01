from __future__ import annotations

import json
import os
import subprocess
import time
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook

from ..config import get_provider_secret
from ..models import StudyRuntime


ENGLISH_LANGS = ["eng-US", "en-US", "en", "eng"]


def ytdlp_binary() -> str:
    return os.environ.get("TIKTOK_RESEARCH_YTDLP", "yt-dlp")


def get_urls_from_workbook(path: Path) -> list[tuple[str, str]]:
    workbook = load_workbook(path)
    try:
        sheet = workbook["Reviewer_A"]
        return [
            (str(row[0]), str(row[1]))
            for row in sheet.iter_rows(min_row=4, values_only=True)
            if row[0] and row[1]
        ]
    finally:
        workbook.close()


def vtt_to_text(vtt_content: str) -> str:
    seen, clean = set(), []
    for line in vtt_content.splitlines():
        line = line.strip()
        if not line or line.startswith("WEBVTT") or "-->" in line or line.isdigit():
            continue
        if line not in seen:
            seen.add(line)
            clean.append(line)
    return " ".join(clean).strip()


def fetch_captions(url: str, vid_id: str, transcripts_dir: Path) -> tuple[bool, str]:
    tmp_prefix = transcripts_dir / f"_tmp_{vid_id}"
    for lang in ENGLISH_LANGS:
        cmd = [
            ytdlp_binary(),
            "--write-auto-sub",
            "--write-subs",
            "--skip-download",
            "--sub-format",
            "vtt",
            "--sub-langs",
            lang,
            "-o",
            str(tmp_prefix),
            "--quiet",
            "--no-warnings",
            url,
        ]
        try:
            subprocess.run(cmd, capture_output=True, timeout=45, check=False)
        except FileNotFoundError as exc:
            raise RuntimeError("yt-dlp is required for transcript extraction but was not found on PATH.") from exc
        except subprocess.TimeoutExpired:
            continue
        matches = list(transcripts_dir.glob(f"_tmp_{vid_id}.{lang}.vtt"))
        if matches:
            text = vtt_to_text(matches[0].read_text(encoding="utf-8", errors="replace"))
            matches[0].unlink(missing_ok=True)
            if text:
                return True, text
    return False, ""


def fetch_whisper(
    runtime: StudyRuntime,
    *,
    url: str,
    vid_id: str,
) -> tuple[bool, str]:
    api_key = get_provider_secret(runtime, "openai", "api_key_env")
    if not api_key:
        raise RuntimeError(
            "Whisper fallback is not configured. Set the env var named in providers.openai.api_key_env."
        )
    cache_dir = runtime.paths.cache_dir / "audio"
    cache_dir.mkdir(parents=True, exist_ok=True)
    cached = list(cache_dir.glob(f"{vid_id}.*"))
    video_path = cached[0] if cached else None
    if video_path is None:
        cmd = [
            ytdlp_binary(),
            "-S",
            "+size",
            "-o",
            str(cache_dir / f"{vid_id}.%(ext)s"),
            "--quiet",
            "--no-warnings",
            url,
        ]
        try:
            result = subprocess.run(cmd, capture_output=True, timeout=180, check=False)
        except FileNotFoundError as exc:
            raise RuntimeError("yt-dlp is required for Whisper fallback but was not found on PATH.") from exc
        except subprocess.TimeoutExpired:
            return False, ""
        if result.returncode != 0:
            return False, ""
        downloaded = list(cache_dir.glob(f"{vid_id}.*"))
        if not downloaded:
            return False, ""
        video_path = downloaded[0]
    with video_path.open("rb") as handle:
        response = requests.post(
            "https://api.openai.com/v1/audio/transcriptions",
            headers={"Authorization": f"Bearer {api_key}"},
            files={"file": (video_path.name, handle, "video/mp4")},
            data={"model": runtime.raw_config["providers"]["openai"].get("whisper_model", "whisper-1"), "language": "en"},
            timeout=120,
        )
    if response.status_code != 200:
        return False, ""
    text = response.json().get("text", "")
    return bool(text), text


def transcribe_workspace(
    runtime: StudyRuntime,
    *,
    use_whisper: bool = False,
    whisper_only: bool = False,
) -> dict[str, Any]:
    runtime.paths.transcripts_dir.mkdir(parents=True, exist_ok=True)
    index_path = runtime.paths.transcripts_dir / "index.json"
    index = json.loads(index_path.read_text(encoding="utf-8")) if index_path.exists() else {}
    videos = get_urls_from_workbook(runtime.paths.workbook_path)
    caption_ok = 0
    whisper_ok = 0
    failed: list[str] = []
    for video_id, url in videos:
        transcript_path = runtime.paths.transcripts_dir / f"{video_id}.txt"
        if video_id in index and index[video_id].get("status") in {"caption", "whisper", "demo"} and transcript_path.exists():
            continue
        ok = False
        text = ""
        method = "none"
        if not whisper_only:
            ok, text = fetch_captions(url, video_id, runtime.paths.transcripts_dir)
            if ok:
                method = "caption"
                caption_ok += 1
        if not ok and (use_whisper or whisper_only):
            ok, text = fetch_whisper(runtime, url=url, vid_id=video_id)
            if ok:
                method = "whisper"
                whisper_ok += 1
        if ok:
            transcript_path.write_text(text, encoding="utf-8")
            index[video_id] = {"url": url, "status": method, "chars": len(text)}
        else:
            failed.append(video_id)
            index[video_id] = {"url": url, "status": "failed"}
        index_path.write_text(json.dumps(index, indent=2) + "\n", encoding="utf-8")
        time.sleep(0.3)
    return {
        "transcripts_dir": str(runtime.paths.transcripts_dir),
        "caption_ok": caption_ok,
        "whisper_ok": whisper_ok,
        "failed": failed,
        "total": len(videos),
    }

