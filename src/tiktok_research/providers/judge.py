from __future__ import annotations

import json
import re
import time
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook

from ..config import get_provider_secret
from ..models import StudyRuntime


CLAUDE_MODEL_DEFAULT = "claude-sonnet-4-6"
GPT_MODEL_DEFAULT = "gpt-5.2-2025-12-11"
MIN_TRANSCRIPT_CHARS = 40


SYSTEM_PROMPT = """You are an expert medical content analyst specializing in urogynecology and urinary incontinence (UI). You are part of a research team conducting a peer-reviewed content analysis of short-form health videos.

Return only valid JSON."""


RUBRIC_PROMPT = """Score the following short-form video transcript for a study on urinary incontinence misinformation.

Video ID: {video_id}
Creator handle: {creator_handle}
Description: {description}

Transcript:
{transcript}

Return JSON with these fields:
{{
  "video_id": "{video_id}",
  "sufficient_content": true,
  "dq1": 0, "dq2": 0, "dq3": 0, "dq4": 0, "dq5": 0,
  "dq6": 0, "dq7": 0, "dq8": 0, "dq9": 0, "dq10": 0,
  "dq11": 0, "dq12": 0, "dq13": 0, "dq14": 0, "dq15": 0,
  "dq16": 0,
  "discern_total": 0,
  "misinformation_score": 0,
  "misinformation_types": [],
  "clinical_domain": [],
  "primary_incontinence_type": "",
  "content_category": "",
  "treatments_mentioned": [],
  "creator_type_inferred": "",
  "key_claims": [],
  "reasoning": ""
}}
"""


def load_metadata(workbook_path: Path) -> dict[str, dict[str, str]]:
    workbook = load_workbook(workbook_path)
    try:
        sheet = workbook["Reviewer_A"]
        meta = {}
        for row in sheet.iter_rows(min_row=4, values_only=True):
            if not row[0]:
                continue
            meta[str(row[0])] = {
                "url": str(row[1] or ""),
                "handle": str(row[7] or "unknown"),
                "description": "",
            }
        return meta
    finally:
        workbook.close()


def load_description_lookup(metadata_json_path: Path) -> dict[str, str]:
    if not metadata_json_path.exists():
        return {}
    payload = json.loads(metadata_json_path.read_text(encoding="utf-8"))
    return {str(item["id"]): str(item.get("video_description") or "") for item in payload if item.get("id")}


def _parse_json(text: str) -> dict[str, Any] | None:
    text = re.sub(r"^```(?:json)?\n?", "", text.strip())
    text = re.sub(r"\n?```$", "", text.strip())
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        return None
    dq_keys = [f"dq{i}" for i in range(1, 17)]
    if all(key in data for key in dq_keys):
        data["discern_total"] = sum(int(data[key] or 0) for key in dq_keys)
    return data


def call_openai(prompt: str, runtime: StudyRuntime) -> dict[str, Any] | None:
    api_key = get_provider_secret(runtime, "openai", "api_key_env")
    if not api_key:
        return None
    model = runtime.raw_config["providers"]["openai"].get("judge_model", GPT_MODEL_DEFAULT)
    response = requests.post(
        "https://api.openai.com/v1/chat/completions",
        headers={
            "Authorization": f"Bearer {api_key}",
            "content-type": "application/json",
        },
        json={
            "model": model,
            "max_completion_tokens": 1500,
            "temperature": 0,
            "messages": [
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": prompt},
            ],
        },
        timeout=90,
    )
    if response.status_code != 200:
        return None
    return _parse_json(response.json()["choices"][0]["message"]["content"])


def call_anthropic(prompt: str, runtime: StudyRuntime) -> dict[str, Any] | None:
    api_key = get_provider_secret(runtime, "anthropic", "api_key_env")
    if not api_key:
        return None
    model = runtime.raw_config["providers"]["anthropic"].get("judge_model", CLAUDE_MODEL_DEFAULT)
    response = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        json={
            "model": model,
            "max_tokens": 1500,
            "temperature": 0,
            "system": SYSTEM_PROMPT,
            "messages": [{"role": "user", "content": prompt}],
        },
        timeout=90,
    )
    if response.status_code != 200:
        return None
    text = response.json()["content"][0]["text"].strip()
    return _parse_json(text)


def score_workspace(runtime: StudyRuntime, *, model: str = "all") -> dict[str, Any]:
    metadata = load_metadata(runtime.paths.workbook_path)
    descriptions = load_description_lookup(runtime.paths.metadata_json_path)
    runtime.paths.llm_scores_dir.mkdir(parents=True, exist_ok=True)
    results = {"gpt": 0, "claude": 0, "skipped": 0}
    for transcript_path in sorted(runtime.paths.transcripts_dir.glob("VID*.txt")):
        video_id = transcript_path.stem
        transcript = transcript_path.read_text(encoding="utf-8").strip()
        if len(transcript) < MIN_TRANSCRIPT_CHARS:
            results["skipped"] += 1
            continue
        meta = metadata.get(video_id, {"handle": "unknown", "description": "", "url": ""})
        numeric_id = re.search(r"/video/(\d+)", meta["url"])
        description = descriptions.get(numeric_id.group(1) if numeric_id else "", meta["description"])
        prompt = RUBRIC_PROMPT.format(
            video_id=video_id,
            creator_handle=meta["handle"],
            description=description,
            transcript=transcript,
        )
        output_dir = runtime.paths.llm_scores_dir / video_id
        output_dir.mkdir(parents=True, exist_ok=True)
        if model in {"all", "gpt"}:
            gpt = call_openai(prompt, runtime)
            if gpt is not None:
                (output_dir / "gpt.json").write_text(json.dumps(gpt, indent=2) + "\n", encoding="utf-8")
                results["gpt"] += 1
        if model in {"all", "claude"}:
            claude = call_anthropic(prompt, runtime)
            if claude is not None:
                (output_dir / "claude.json").write_text(json.dumps(claude, indent=2) + "\n", encoding="utf-8")
                results["claude"] += 1
        time.sleep(0.4)
    return results
