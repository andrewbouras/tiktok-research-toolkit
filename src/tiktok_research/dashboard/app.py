from __future__ import annotations

import csv
import io
import json
import os
import re
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

from flask import (
    Flask,
    abort,
    g,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    url_for,
)
from openpyxl import load_workbook

from ..config import load_runtime
from ..form_schema import DISCERN_FIELDS, EXPORT_FIELD_ORDER, FIELD_INDEX, FIELD_ORDER, FIELD_SECTIONS
from ..models import StudyRuntime


APP_DIR = Path(__file__).resolve().parent
PRIMARY_VIDEO_TABLE = "videos"
CALIBRATION_VIDEO_TABLE = "calibration_videos"
PRIMARY_RATING_TABLE = "ratings"
CALIBRATION_RATING_TABLE = "calibration_ratings"
RATING_TABLES = {PRIMARY_RATING_TABLE, CALIBRATION_RATING_TABLE}
VIDEO_TABLES = {PRIMARY_VIDEO_TABLE, CALIBRATION_VIDEO_TABLE}
CALIBRATION_VIDEO_COUNT = 10
CALIBRATION_DISCERN_SPREAD_THRESHOLD = 5
CALIBRATION_VIDEO_ID_PREFIX = "CAL"
CALIBRATION_MIN_RELEVANCE_SCORE = 3
CALIBRATION_KEYWORD_WEIGHTS = {
    "urinary incontinence": 8,
    "stress incontinence": 8,
    "urge incontinence": 8,
    "bladder leakage": 8,
    "urine leakage": 8,
    "leaking urine": 8,
    "overactive bladder": 7,
    "incontinence": 7,
    "leak when": 5,
    "pelvic floor": 4,
    "bladder": 3,
    "kegel": 3,
    "postpartum": 2,
    "pee": 2,
    "peeing": 2,
    "leak": 2,
    "prolapse": 1,
}
DEFAULT_RESEARCHERS: list[tuple[str, str]] = []
STUDY_TITLE = "TikTok Research Toolkit"
BACKUP_MIN_INTERVAL_SECONDS = 300
BACKUP_MAX_FILES = 96
DATA_DIR = Path(".")
DATABASE_PATH = Path("research_dashboard.sqlite3")
BACKUP_DIR = Path("backups")
WORKBOOK_PATH = Path("coding_template.xlsx")
COLLECTED_VIDEOS_PATH = Path("collected_videos.json")
REPLACEMENT_CANDIDATES_PATH = Path("replacement_candidates.json")
TRANSCRIPTS_DIR = Path("transcripts")
CURRENT_RUNTIME: StudyRuntime | None = None


def configure_runtime(runtime_or_root: StudyRuntime | Path | str) -> StudyRuntime:
    runtime = runtime_or_root if isinstance(runtime_or_root, StudyRuntime) else load_runtime(Path(runtime_or_root))
    global CURRENT_RUNTIME
    global DEFAULT_RESEARCHERS
    global STUDY_TITLE
    global BACKUP_MIN_INTERVAL_SECONDS
    global BACKUP_MAX_FILES
    global DATA_DIR
    global DATABASE_PATH
    global BACKUP_DIR
    global WORKBOOK_PATH
    global COLLECTED_VIDEOS_PATH
    global REPLACEMENT_CANDIDATES_PATH
    global TRANSCRIPTS_DIR

    CURRENT_RUNTIME = runtime
    DEFAULT_RESEARCHERS = [(item.slug, item.display_name) for item in runtime.researchers]
    STUDY_TITLE = str(runtime.raw_config["study"].get("name") or "TikTok Research Toolkit")
    BACKUP_MIN_INTERVAL_SECONDS = int(runtime.raw_config["study"].get("backup_min_interval_seconds", 300))
    BACKUP_MAX_FILES = int(runtime.raw_config["study"].get("backup_max_files", 96))
    DATA_DIR = runtime.paths.data_dir
    DATABASE_PATH = runtime.paths.database_path
    BACKUP_DIR = runtime.paths.backups_dir
    WORKBOOK_PATH = runtime.paths.workbook_path
    COLLECTED_VIDEOS_PATH = runtime.paths.metadata_json_path
    REPLACEMENT_CANDIDATES_PATH = runtime.paths.replacement_candidates_path
    TRANSCRIPTS_DIR = runtime.paths.transcripts_dir
    return runtime


def current_runtime() -> StudyRuntime:
    if CURRENT_RUNTIME is None:
        workspace = os.environ.get("TIKTOK_RESEARCH_WORKSPACE")
        if not workspace:
            raise RuntimeError(
                "No workspace configured. Pass a workspace to create_app() or set TIKTOK_RESEARCH_WORKSPACE."
            )
        return configure_runtime(Path(workspace))
    return CURRENT_RUNTIME


def resolve_rating_table(table: str) -> str:
    if table not in RATING_TABLES:
        raise ValueError(f"Unsupported ratings table: {table}")
    return table


def resolve_video_table(table: str) -> str:
    if table not in VIDEO_TABLES:
        raise ValueError(f"Unsupported video table: {table}")
    return table


def compact_sheet_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())


def create_app(workspace_root: StudyRuntime | Path | str | None = None) -> Flask:
    runtime = current_runtime() if workspace_root is None else configure_runtime(workspace_root)
    app = Flask(__name__, template_folder="templates", static_folder="static")
    app.config["DATABASE"] = str(DATABASE_PATH)
    app.config["WORKSPACE_ROOT"] = str(runtime.paths.root)
    app.config["STUDY_TITLE"] = STUDY_TITLE

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ensure_database()

    @app.teardown_appcontext
    def close_connection(exception: Exception | None) -> None:
        db = g.pop("db", None)
        if db is not None:
            db.close()

    @app.template_filter("compact_number")
    def compact_number(value: Any) -> str:
        try:
            number = int(value)
        except (TypeError, ValueError):
            return "0"
        if number >= 1_000_000:
            return f"{number / 1_000_000:.1f}M"
        if number >= 1_000:
            return f"{number / 1_000:.1f}K"
        return str(number)

    @app.template_filter("pretty_dt")
    def pretty_dt(value: str | None) -> str:
        if not value:
            return "No activity yet"
        try:
            return datetime.fromisoformat(value).strftime("%b %d, %Y %I:%M %p")
        except ValueError:
            return value

    @app.context_processor
    def inject_navigation() -> dict[str, Any]:
        active_slug = None
        if request.view_args:
            active_slug = request.view_args.get("slug")
        if not active_slug:
            active_slug = request.cookies.get("preferred_researcher")
        endpoint = request.endpoint or ""
        nav_active_page = "overview"
        if endpoint.startswith("calibration"):
            nav_active_page = "calibration"
        return {
            "study_title": STUDY_TITLE,
            "workspace_root": str(runtime.paths.root),
            "nav_researchers": fetch_researchers(get_db()),
            "nav_active_slug": active_slug,
            "nav_active_page": nav_active_page,
        }

    @app.route("/")
    def dashboard() -> str:
        conn = get_db()
        researchers = fetch_researchers(conn)
        total_videos = conn.execute("SELECT COUNT(*) AS count FROM videos").fetchone()["count"]
        ratings = fetch_all_ratings(conn)
        total_possible = total_videos * len(researchers)
        completed_count = sum(1 for row in ratings if row["is_complete"])
        in_progress_count = sum(
            1 for row in ratings if not row["is_complete"] and row["progress_pct"] > 0
        )
        flagged_count = sum(1 for row in ratings if row["flag_for_review"])
        overall_progress = round((completed_count / total_possible) * 100, 1) if total_possible else 0

        misinfo_counts = {str(score): 0 for score in range(1, 6)}
        discern_values: list[int] = []
        for row in ratings:
            responses = row["responses"]
            misinfo_score = responses.get("B3_Misinfo_Score")
            if misinfo_score in misinfo_counts:
                misinfo_counts[misinfo_score] += 1
            discern_total = compute_discern_total(responses)
            if row["is_complete"] and discern_total is not None:
                discern_values.append(discern_total)

        researcher_cards = []
        for researcher in researchers:
            stats = build_researcher_stats(conn, researcher["slug"])
            researcher_cards.append(
                {
                    "display_name": researcher["display_name"],
                    "slug": researcher["slug"],
                    "deletable": researcher_is_deletable(researcher["slug"]),
                    **stats,
                }
            )

        recent_activity = conn.execute(
            """
            SELECT r.updated_at, r.video_id, rs.display_name
            FROM ratings r
            JOIN researchers rs ON rs.id = r.researcher_id
            WHERE r.updated_at IS NOT NULL
            ORDER BY r.updated_at DESC
            LIMIT 8
            """
        ).fetchall()

        preferred_slug = request.cookies.get("preferred_researcher")
        return render_template(
            "dashboard.html",
            total_videos=total_videos,
            total_researchers=len(researchers),
            completed_count=completed_count,
            in_progress_count=in_progress_count,
            flagged_count=flagged_count,
            overall_progress=overall_progress,
            avg_discern=round(sum(discern_values) / len(discern_values), 1) if discern_values else None,
            misinfo_counts=misinfo_counts,
            researcher_cards=researcher_cards,
            recent_activity=recent_activity,
            preferred_slug=preferred_slug,
        )

    @app.route("/calibration")
    def calibration_dashboard() -> str:
        conn = get_db()
        researchers = fetch_researchers(conn)
        calibration_video_ids = fetch_calibration_video_ids(conn)
        ratings = fetch_all_ratings(conn, table=CALIBRATION_RATING_TABLE, video_ids=calibration_video_ids)
        total_possible = len(calibration_video_ids) * len(researchers)
        completed_count = sum(1 for row in ratings if row["is_complete"])
        in_progress_count = sum(
            1 for row in ratings if not row["is_complete"] and row["progress_pct"] > 0
        )
        overall_progress = round((completed_count / total_possible) * 100, 1) if total_possible else 0

        researcher_cards = []
        for researcher in researchers:
            stats = build_researcher_stats(
                conn,
                researcher["slug"],
                table=CALIBRATION_RATING_TABLE,
                video_ids=calibration_video_ids,
                video_table=CALIBRATION_VIDEO_TABLE,
            )
            researcher_cards.append({"display_name": researcher["display_name"], "slug": researcher["slug"], **stats})

        comparison_rows = build_calibration_comparison_rows(conn, calibration_video_ids, researchers)
        preferred_slug = request.cookies.get("preferred_researcher")
        return render_template(
            "calibration.html",
            total_videos=len(calibration_video_ids),
            total_researchers=len(researchers),
            completed_count=completed_count,
            in_progress_count=in_progress_count,
            overall_progress=overall_progress,
            researcher_cards=researcher_cards,
            preferred_slug=preferred_slug,
            comparison_rows=comparison_rows,
            videos_needing_discussion=sum(1 for row in comparison_rows if row["needs_discussion"]),
        )

    @app.post("/researchers/<slug>/delete")
    def delete_researcher(slug: str) -> Any:
        conn = get_db()
        researcher = fetch_researcher_by_slug(conn, slug)
        if researcher is None:
            abort(404)
        if not researcher_is_deletable(researcher["slug"]):
            abort(403)
        conn.execute("DELETE FROM ratings WHERE researcher_id = ?", (researcher["id"],))
        conn.execute("DELETE FROM calibration_ratings WHERE researcher_id = ?", (researcher["id"],))
        conn.execute("DELETE FROM researchers WHERE id = ?", (researcher["id"],))
        conn.commit()
        response = make_response(redirect(url_for("dashboard")))
        preferred_slug = request.cookies.get("preferred_researcher")
        if preferred_slug == researcher["slug"]:
            response.delete_cookie("preferred_researcher")
        return response

    @app.route("/researchers", methods=["POST"])
    def create_researcher() -> Any:
        name = (request.form.get("display_name") or "").strip()
        if not name:
            return redirect(url_for("dashboard"))
        conn = get_db()
        existing = conn.execute(
            "SELECT slug FROM researchers WHERE lower(display_name) = lower(?)",
            (name,),
        ).fetchone()
        if existing is not None:
            response = make_response(redirect(url_for("researcher_start", slug=existing["slug"])))
            response.set_cookie("preferred_researcher", existing["slug"], max_age=60 * 60 * 24 * 180)
            return response
        slug = unique_slug(conn, slugify(name))
        now = utcnow()
        conn.execute(
            """
            INSERT INTO researchers (slug, display_name, created_at, updated_at)
            VALUES (?, ?, ?, ?)
            """,
            (slug, name, now, now),
        )
        conn.commit()
        response = make_response(redirect(url_for("researcher_start", slug=slug)))
        response.set_cookie("preferred_researcher", slug, max_age=60 * 60 * 24 * 180)
        return response

    @app.route("/calibration/<slug>")
    def calibration_start(slug: str) -> Any:
        conn = get_db()
        researcher = fetch_researcher_by_slug(conn, slug)
        if researcher is None:
            abort(404)
        calibration_video_ids = fetch_calibration_video_ids(conn)
        videos = fetch_queue_rows(
            conn,
            researcher["id"],
            table=CALIBRATION_RATING_TABLE,
            video_ids=calibration_video_ids,
            video_table=CALIBRATION_VIDEO_TABLE,
        )
        resume_video = pick_resume_video(videos)
        if resume_video is None:
            return redirect(url_for("calibration_dashboard"))
        response = make_response(
            redirect(url_for("calibration_workspace", slug=researcher["slug"], video_id=resume_video["id"]))
        )
        response.set_cookie("preferred_researcher", researcher["slug"], max_age=60 * 60 * 24 * 180)
        return response

    @app.route("/calibration/<slug>/queue")
    def calibration_queue(slug: str) -> Any:
        conn = get_db()
        researcher = fetch_researcher_by_slug(conn, slug)
        if researcher is None:
            abort(404)
        calibration_video_ids = fetch_calibration_video_ids(conn)
        videos = fetch_queue_rows(
            conn,
            researcher["id"],
            table=CALIBRATION_RATING_TABLE,
            video_ids=calibration_video_ids,
            video_table=CALIBRATION_VIDEO_TABLE,
        )
        for video in videos:
            video["workspace_url"] = url_for("calibration_workspace", slug=researcher["slug"], video_id=video["id"])
        stats = build_researcher_stats(
            conn,
            slug,
            table=CALIBRATION_RATING_TABLE,
            video_ids=calibration_video_ids,
            video_table=CALIBRATION_VIDEO_TABLE,
        )
        resume_video = pick_resume_video(videos)
        response = make_response(
            render_template(
                "researcher.html",
                researcher=researcher,
                videos=videos,
                stats=stats,
                resume_video=resume_video,
                can_delete_researcher=researcher_is_deletable(researcher["slug"]),
                page_title=f"{researcher['display_name']} Calibration",
                queue_eyebrow="Calibration set",
                queue_heading="Calibration videos",
                resume_url=url_for("calibration_start", slug=researcher["slug"]),
                resume_label="Continue calibration",
                export_url=url_for("export_calibration_researcher_csv", slug=researcher["slug"]),
                export_label="Export calibration CSV",
                search_placeholder="Search calibration videos",
                queue_open_not_started_label="Open calibration",
                queue_open_in_progress_label="Resume calibration",
                queue_open_complete_label="Review calibration",
            )
        )
        response.set_cookie("preferred_researcher", researcher["slug"], max_age=60 * 60 * 24 * 180)
        return response

    @app.route("/researchers/<slug>")
    def researcher_start(slug: str) -> Any:
        conn = get_db()
        researcher = fetch_researcher_by_slug(conn, slug)
        if researcher is None:
            abort(404)
        videos = fetch_queue_rows(conn, researcher["id"])
        resume_video = pick_resume_video(videos)
        if resume_video is None:
            return redirect(url_for("researcher_queue", slug=researcher["slug"]))
        response = make_response(
            redirect(url_for("video_workspace", slug=researcher["slug"], video_id=resume_video["id"]))
        )
        response.set_cookie("preferred_researcher", researcher["slug"], max_age=60 * 60 * 24 * 180)
        return response

    @app.route("/researchers/<slug>/queue")
    def researcher_queue(slug: str) -> Any:
        conn = get_db()
        researcher = fetch_researcher_by_slug(conn, slug)
        if researcher is None:
            abort(404)
        videos = fetch_queue_rows(conn, researcher["id"])
        for video in videos:
            video["workspace_url"] = url_for("video_workspace", slug=researcher["slug"], video_id=video["id"])
        stats = build_researcher_stats(conn, slug)
        resume_video = pick_resume_video(videos)
        response = make_response(
            render_template(
                "researcher.html",
                researcher=researcher,
                videos=videos,
                stats=stats,
                resume_video=resume_video,
                can_delete_researcher=researcher_is_deletable(researcher["slug"]),
                resume_url=url_for("researcher_start", slug=researcher["slug"]),
                export_url=url_for("export_researcher_csv", slug=researcher["slug"]),
            )
        )
        response.set_cookie("preferred_researcher", researcher["slug"], max_age=60 * 60 * 24 * 180)
        return response

    @app.route("/researchers/<slug>/videos/<video_id>")
    def video_workspace(slug: str, video_id: str) -> Any:
        return render_video_workspace(slug, video_id, calibration=False)

    @app.route("/calibration/<slug>/videos/<video_id>")
    def calibration_workspace(slug: str, video_id: str) -> Any:
        return render_video_workspace(slug, video_id, calibration=True)

    @app.post("/api/researchers/<slug>/videos/<video_id>")
    def save_rating(slug: str, video_id: str) -> Any:
        return save_rating_endpoint(slug, video_id, calibration=False)

    @app.post("/api/calibration/<slug>/videos/<video_id>")
    def save_calibration_rating(slug: str, video_id: str) -> Any:
        return save_rating_endpoint(slug, video_id, calibration=True)

    @app.route("/exports/all.csv")
    def export_all_csv() -> Any:
        conn = get_db()
        rows: list[dict[str, Any]] = []
        for researcher in fetch_researchers(conn):
            rows.extend(export_rows_for_researcher(conn, researcher["id"], researcher["display_name"]))
        return csv_response(rows, "tiktok_research_all_coders.csv")

    @app.route("/exports/<slug>.csv")
    def export_researcher_csv(slug: str) -> Any:
        conn = get_db()
        researcher = fetch_researcher_by_slug(conn, slug)
        if researcher is None:
            abort(404)
        rows = export_rows_for_researcher(conn, researcher["id"], researcher["display_name"])
        return csv_response(rows, f"{slug}_coding_export.csv")

    @app.route("/exports/calibration/all.csv")
    def export_calibration_all_csv() -> Any:
        conn = get_db()
        calibration_video_ids = fetch_calibration_video_ids(conn)
        rows: list[dict[str, Any]] = []
        for researcher in fetch_researchers(conn):
            rows.extend(
                export_rows_for_researcher(
                    conn,
                    researcher["id"],
                    researcher["display_name"],
                    table=CALIBRATION_RATING_TABLE,
                    video_ids=calibration_video_ids,
                    video_table=CALIBRATION_VIDEO_TABLE,
                )
            )
        return csv_response(rows, "tiktok_research_calibration_all_coders.csv")

    @app.route("/exports/calibration/<slug>.csv")
    def export_calibration_researcher_csv(slug: str) -> Any:
        conn = get_db()
        researcher = fetch_researcher_by_slug(conn, slug)
        if researcher is None:
            abort(404)
        calibration_video_ids = fetch_calibration_video_ids(conn)
        rows = export_rows_for_researcher(
            conn,
            researcher["id"],
            researcher["display_name"],
            table=CALIBRATION_RATING_TABLE,
            video_ids=calibration_video_ids,
            video_table=CALIBRATION_VIDEO_TABLE,
        )
        return csv_response(rows, f"{slug}_calibration_export.csv")

    @app.route("/healthz")
    def healthz() -> Any:
        provider_summary = {}
        for provider_name, provider in runtime.raw_config.get("providers", {}).items():
            if isinstance(provider, dict):
                provider_summary[provider_name] = {"enabled": bool(provider.get("enabled"))}
        return jsonify(
            {
                "ok": True,
                "time": utcnow(),
                "study": STUDY_TITLE,
                "workspace_root": str(runtime.paths.root),
                "providers": provider_summary,
            }
        )

    return app


def render_video_workspace(slug: str, video_id: str, *, calibration: bool) -> Any:
    conn = get_db()
    researcher = fetch_researcher_by_slug(conn, slug)
    if researcher is None:
        abort(404)

    video_table = CALIBRATION_VIDEO_TABLE if calibration else PRIMARY_VIDEO_TABLE
    rating_table = CALIBRATION_RATING_TABLE if calibration else PRIMARY_RATING_TABLE
    video_ids = fetch_calibration_video_ids(conn) if calibration else None
    if calibration and video_id not in (video_ids or []):
        abort(404)

    video = conn.execute(f"SELECT * FROM {video_table} WHERE id = ?", (video_id,)).fetchone()
    if video is None:
        abort(404)
    rating = fetch_or_create_rating(conn, researcher["id"], video_id, table=rating_table)
    queue_rows = fetch_queue_rows(
        conn,
        researcher["id"],
        table=rating_table,
        video_ids=video_ids,
        video_table=video_table,
    )
    ordered_video_ids = [row["id"] for row in queue_rows]
    current_index = ordered_video_ids.index(video_id)
    previous_video_id = ordered_video_ids[current_index - 1] if current_index > 0 else None
    next_video_id = ordered_video_ids[current_index + 1] if current_index < len(ordered_video_ids) - 1 else None
    stats = build_researcher_stats(
        conn,
        slug,
        table=rating_table,
        video_ids=video_ids,
        video_table=video_table,
    )
    response = make_response(
        render_template(
            "video.html",
            researcher=researcher,
            video=video,
            rating=rating,
            field_sections=FIELD_SECTIONS,
            previous_video_id=previous_video_id,
            next_video_id=next_video_id,
            initial_missing=required_field_ids(rating["responses"]),
            api_url=url_for("save_calibration_rating" if calibration else "save_rating", slug=slug, video_id=video_id),
            queue_url=url_for("calibration_queue" if calibration else "researcher_queue", slug=slug),
            start_url=url_for("calibration_start" if calibration else "researcher_start", slug=slug),
            embed_url=build_embed_url(video["tiktok_numeric_id"]),
            current_index=current_index + 1,
            total_videos=len(ordered_video_ids),
            primary_link_label="Calibration video" if calibration else "Assigned video",
            queue_link_label="All calibration videos" if calibration else "All videos",
            workspace_eyebrow="Calibration workspace" if calibration else "Focused coding workspace",
            continue_button_label="Continue calibration" if calibration else "Continue assigned work",
            overall_progress_label="Calibration set" if calibration else "All assigned videos",
            return_later_url=url_for("calibration_dashboard" if calibration else "dashboard"),
            return_later_label="Save and return to calibration" if calibration else "Save and return later",
            return_to_flow_label="Return to calibration set" if calibration else "Return to assigned work",
            mode_note=(
                "These videos are for group calibration only and do not overlap with the main coding set."
                if calibration
                else ""
            ),
            researcher_progress_pct=stats.get("progress_pct", 0),
            researcher_completed_count=stats.get("completed", 0),
            researcher_total_count=stats.get("total", len(ordered_video_ids)),
            discern_total=compute_discern_total(rating["responses"]),
        )
    )
    response.set_cookie("preferred_researcher", researcher["slug"], max_age=60 * 60 * 24 * 180)
    return response


def save_rating_endpoint(slug: str, video_id: str, *, calibration: bool) -> Any:
    conn = get_db()
    researcher = fetch_researcher_by_slug(conn, slug)
    if researcher is None:
        abort(404)
    video_table = CALIBRATION_VIDEO_TABLE if calibration else PRIMARY_VIDEO_TABLE
    rating_table = CALIBRATION_RATING_TABLE if calibration else PRIMARY_RATING_TABLE
    row = conn.execute(f"SELECT id FROM {video_table} WHERE id = ?", (video_id,)).fetchone()
    if row is None:
        abort(404)
    payload = request.get_json(silent=True) or {}
    responses = sanitize_responses(payload.get("responses") or {})
    result = save_rating_responses(conn, researcher["id"], video_id, responses, table=rating_table)
    try:
        write_backup_snapshot(conn, reason="autosave")
    except Exception:
        pass
    return jsonify(result)


def get_db() -> sqlite3.Connection:
    if "db" not in g:
        connection = sqlite3.connect(DATABASE_PATH)
        configure_sqlite_connection(connection)
        g.db = connection
    return g.db


def ensure_database() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DATABASE_PATH)
    configure_sqlite_connection(conn)
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS videos (
            id TEXT PRIMARY KEY,
            tiktok_url TEXT NOT NULL,
            tiktok_numeric_id TEXT,
            views INTEGER,
            likes INTEGER,
            comments INTEGER,
            shares INTEGER,
            post_date TEXT,
            creator_handle TEXT,
            description TEXT,
            transcript TEXT,
            source_row INTEGER
        );
        CREATE TABLE IF NOT EXISTS calibration_videos (
            id TEXT PRIMARY KEY,
            tiktok_url TEXT NOT NULL,
            tiktok_numeric_id TEXT,
            views INTEGER,
            likes INTEGER,
            comments INTEGER,
            shares INTEGER,
            post_date TEXT,
            creator_handle TEXT,
            description TEXT,
            transcript TEXT,
            source_row INTEGER
        );
        CREATE TABLE IF NOT EXISTS researchers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            slug TEXT UNIQUE NOT NULL,
            display_name TEXT UNIQUE NOT NULL,
            created_at TEXT,
            updated_at TEXT
        );
        CREATE TABLE IF NOT EXISTS ratings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            researcher_id INTEGER NOT NULL,
            video_id TEXT NOT NULL,
            responses_json TEXT NOT NULL DEFAULT '{}',
            progress_pct REAL NOT NULL DEFAULT 0,
            is_complete INTEGER NOT NULL DEFAULT 0,
            flag_for_review INTEGER NOT NULL DEFAULT 0,
            started_at TEXT,
            updated_at TEXT,
            completed_at TEXT,
            UNIQUE(researcher_id, video_id),
            FOREIGN KEY(researcher_id) REFERENCES researchers(id),
            FOREIGN KEY(video_id) REFERENCES videos(id)
        );
        CREATE TABLE IF NOT EXISTS calibration_ratings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            researcher_id INTEGER NOT NULL,
            video_id TEXT NOT NULL,
            responses_json TEXT NOT NULL DEFAULT '{}',
            progress_pct REAL NOT NULL DEFAULT 0,
            is_complete INTEGER NOT NULL DEFAULT 0,
            flag_for_review INTEGER NOT NULL DEFAULT 0,
            started_at TEXT,
            updated_at TEXT,
            completed_at TEXT,
            UNIQUE(researcher_id, video_id),
            FOREIGN KEY(researcher_id) REFERENCES researchers(id),
            FOREIGN KEY(video_id) REFERENCES calibration_videos(id)
        );
        """
    )
    conn.commit()
    if conn.execute("SELECT COUNT(*) AS count FROM videos").fetchone()["count"] == 0:
        seed_videos(conn)
    sync_default_researchers(conn)
    ensure_calibration_dataset(conn)
    conn.close()


def seed_videos(conn: sqlite3.Connection) -> None:
    if not WORKBOOK_PATH.exists():
        return
    description_lookup = load_collected_videos_lookup()
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    try:
        sheet = workbook["Reviewer_A"]
        headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
        rows_to_insert = []
        for row_index in range(4, sheet.max_row + 1):
            values = [sheet.cell(row_index, column).value for column in range(1, sheet.max_column + 1)]
            row = dict(zip(headers, values))
            video_id = row.get("Video_ID")
            if not video_id:
                continue
            tiktok_url = row.get("TikTok_URL") or ""
            tiktok_numeric_id = extract_tiktok_numeric_id(tiktok_url)
            transcript_path = TRANSCRIPTS_DIR / f"{video_id}.txt"
            transcript = transcript_path.read_text(encoding="utf-8").strip() if transcript_path.exists() else ""
            description = ""
            if tiktok_numeric_id and tiktok_numeric_id in description_lookup:
                description = description_lookup[tiktok_numeric_id].get("video_description") or ""
            rows_to_insert.append(
                (
                    video_id,
                    tiktok_url,
                    tiktok_numeric_id,
                    to_int(row.get("Views")),
                    to_int(row.get("Likes")),
                    to_int(row.get("Comments")),
                    to_int(row.get("Shares")),
                    row.get("Post_Date"),
                    row.get("Creator_Handle"),
                    description,
                    transcript,
                    row_index,
                )
            )
    finally:
        workbook.close()
    if rows_to_insert:
        conn.executemany(
            """
            INSERT INTO videos (
                id, tiktok_url, tiktok_numeric_id, views, likes, comments, shares,
                post_date, creator_handle, description, transcript, source_row
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows_to_insert,
        )
        conn.commit()


def load_collected_videos_lookup() -> dict[str, dict[str, Any]]:
    if not COLLECTED_VIDEOS_PATH.exists():
        return {}
    items = json.loads(COLLECTED_VIDEOS_PATH.read_text(encoding="utf-8"))
    lookup = {}
    for item in items:
        item_id = str(item.get("id") or "")
        if item_id:
            lookup[item_id] = item
    return lookup


def load_replacement_candidates() -> list[dict[str, Any]]:
    if not REPLACEMENT_CANDIDATES_PATH.exists():
        return []
    try:
        items = json.loads(REPLACEMENT_CANDIDATES_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return []
    return items if isinstance(items, list) else []


def table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type = 'table' AND name = ?",
        (table_name,),
    ).fetchone()
    return row is not None


def ensure_calibration_dataset(conn: sqlite3.Connection) -> None:
    rebuild_calibration_rating_table_if_needed(conn)
    if calibration_dataset_needs_refresh(conn):
        reseed_calibration_videos(conn)


def rebuild_calibration_rating_table_if_needed(conn: sqlite3.Connection) -> None:
    foreign_keys = conn.execute(f"PRAGMA foreign_key_list({CALIBRATION_RATING_TABLE})").fetchall()
    video_targets = [row["table"] for row in foreign_keys if row["from"] == "video_id"]
    if video_targets == [CALIBRATION_VIDEO_TABLE]:
        return
    if table_exists(conn, CALIBRATION_RATING_TABLE):
        conn.execute(f"DROP TABLE IF EXISTS {CALIBRATION_RATING_TABLE}")
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {CALIBRATION_RATING_TABLE} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            researcher_id INTEGER NOT NULL,
            video_id TEXT NOT NULL,
            responses_json TEXT NOT NULL DEFAULT '{{}}',
            progress_pct REAL NOT NULL DEFAULT 0,
            is_complete INTEGER NOT NULL DEFAULT 0,
            flag_for_review INTEGER NOT NULL DEFAULT 0,
            started_at TEXT,
            updated_at TEXT,
            completed_at TEXT,
            UNIQUE(researcher_id, video_id),
            FOREIGN KEY(researcher_id) REFERENCES researchers(id),
            FOREIGN KEY(video_id) REFERENCES {CALIBRATION_VIDEO_TABLE}(id)
        )
        """
    )
    conn.commit()


def calibration_dataset_needs_refresh(conn: sqlite3.Connection) -> bool:
    rows = conn.execute(
        f"""
        SELECT id, tiktok_numeric_id
        FROM {CALIBRATION_VIDEO_TABLE}
        ORDER BY source_row ASC
        """
    ).fetchall()
    if len(rows) != CALIBRATION_VIDEO_COUNT:
        return True
    main_numeric_ids = fetch_main_sample_numeric_ids(conn)
    for index, row in enumerate(rows, start=1):
        expected_id = f"{CALIBRATION_VIDEO_ID_PREFIX}{index:03d}"
        if row["id"] != expected_id:
            return True
        numeric_id = str(row["tiktok_numeric_id"] or "").strip()
        if not numeric_id or numeric_id in main_numeric_ids:
            return True
    return False


def fetch_main_sample_numeric_ids(conn: sqlite3.Connection) -> set[str]:
    return {
        str(row["tiktok_numeric_id"]).strip()
        for row in conn.execute(
            """
            SELECT tiktok_numeric_id
            FROM videos
            WHERE tiktok_numeric_id IS NOT NULL
              AND trim(tiktok_numeric_id) != ''
            """
        ).fetchall()
    }


def reseed_calibration_videos(conn: sqlite3.Connection) -> None:
    candidates = select_calibration_candidates(conn)
    conn.execute(f"DELETE FROM {CALIBRATION_RATING_TABLE}")
    conn.execute(f"DELETE FROM {CALIBRATION_VIDEO_TABLE}")
    rows_to_insert = []
    for index, candidate in enumerate(candidates[:CALIBRATION_VIDEO_COUNT], start=1):
        rows_to_insert.append(
            (
                f"{CALIBRATION_VIDEO_ID_PREFIX}{index:03d}",
                build_tiktok_url(candidate["username"], candidate["id"]),
                candidate["id"],
                to_int(candidate.get("view_count")),
                to_int(candidate.get("like_count")),
                to_int(candidate.get("comment_count")),
                to_int(candidate.get("share_count")),
                format_candidate_post_date(candidate.get("create_time")),
                normalize_creator_handle(candidate.get("username")),
                compact_sheet_text(candidate.get("video_description")),
                "",
                index,
            )
        )
    if rows_to_insert:
        conn.executemany(
            f"""
            INSERT INTO {CALIBRATION_VIDEO_TABLE} (
                id, tiktok_url, tiktok_numeric_id, views, likes, comments, shares,
                post_date, creator_handle, description, transcript, source_row
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows_to_insert,
        )
    conn.commit()


def select_calibration_candidates(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    items = load_replacement_candidates()
    if not items:
        return []
    main_numeric_ids = fetch_main_sample_numeric_ids(conn)
    pool: list[dict[str, Any]] = []
    seen_numeric_ids: set[str] = set()
    for item in items:
        numeric_id = str(item.get("id") or "").strip()
        if not numeric_id or numeric_id in seen_numeric_ids or numeric_id in main_numeric_ids:
            continue
        username = str(item.get("username") or "").strip()
        description = compact_sheet_text(item.get("video_description"))
        if not username or not description:
            continue
        score = score_calibration_candidate(item)
        if score < CALIBRATION_MIN_RELEVANCE_SCORE:
            continue
        pool.append({**item, "id": numeric_id, "username": username, "video_description": description, "_relevance_score": score})
        seen_numeric_ids.add(numeric_id)
    pool.sort(
        key=lambda item: (
            -int(item.get("_relevance_score") or 0),
            -to_int(item.get("view_count") or 0) if item.get("view_count") not in (None, "") else 0,
            item["id"],
        )
    )
    selected: list[dict[str, Any]] = []
    used_usernames: set[str] = set()
    used_description_prefixes: set[str] = set()
    for require_unique_username in (True, False):
        for item in pool:
            if item in selected:
                continue
            username_key = item["username"].lower()
            description_key = item["video_description"][:140].lower()
            if require_unique_username and username_key in used_usernames:
                continue
            if description_key in used_description_prefixes:
                continue
            selected.append(item)
            used_usernames.add(username_key)
            used_description_prefixes.add(description_key)
            if len(selected) >= CALIBRATION_VIDEO_COUNT:
                return selected
    return selected


def score_calibration_candidate(item: dict[str, Any]) -> int:
    haystack = " ".join([str(item.get("video_description") or ""), str(item.get("username") or "")]).lower()
    return sum(weight for phrase, weight in CALIBRATION_KEYWORD_WEIGHTS.items() if phrase in haystack)


def fetch_researchers(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute("SELECT * FROM researchers ORDER BY display_name COLLATE NOCASE").fetchall()


def fetch_researcher_by_slug(conn: sqlite3.Connection, slug: str) -> sqlite3.Row | None:
    return conn.execute("SELECT * FROM researchers WHERE slug = ?", (slug,)).fetchone()


def fetch_all_ratings(
    conn: sqlite3.Connection,
    table: str = PRIMARY_RATING_TABLE,
    video_ids: list[str] | None = None,
) -> list[dict[str, Any]]:
    table_name = resolve_rating_table(table)
    rows = conn.execute(
        f"""
        SELECT r.*, rs.display_name
        FROM {table_name} r
        JOIN researchers rs ON rs.id = r.researcher_id
        ORDER BY r.updated_at DESC
        """
    ).fetchall()
    hydrated_rows = [hydrate_rating_row(row) | {"display_name": row["display_name"]} for row in rows]
    if not video_ids:
        return hydrated_rows
    allowed_ids = set(video_ids)
    return [row for row in hydrated_rows if row["video_id"] in allowed_ids]


def fetch_or_create_rating(
    conn: sqlite3.Connection,
    researcher_id: int,
    video_id: str,
    table: str = PRIMARY_RATING_TABLE,
) -> dict[str, Any]:
    table_name = resolve_rating_table(table)
    row = conn.execute(
        f"SELECT * FROM {table_name} WHERE researcher_id = ? AND video_id = ?",
        (researcher_id, video_id),
    ).fetchone()
    if row is None:
        conn.execute(
            f"""
            INSERT INTO {table_name} (
                researcher_id, video_id, responses_json, progress_pct, is_complete,
                flag_for_review, started_at, updated_at, completed_at
            )
            VALUES (?, ?, '{{}}', 0, 0, 0, NULL, NULL, NULL)
            """,
            (researcher_id, video_id),
        )
        conn.commit()
        row = conn.execute(
            f"SELECT * FROM {table_name} WHERE researcher_id = ? AND video_id = ?",
            (researcher_id, video_id),
        ).fetchone()
    return hydrate_rating_row(row)


def fetch_queue_rows(
    conn: sqlite3.Connection,
    researcher_id: int,
    table: str = PRIMARY_RATING_TABLE,
    video_ids: list[str] | None = None,
    video_table: str = PRIMARY_VIDEO_TABLE,
) -> list[dict[str, Any]]:
    table_name = resolve_rating_table(table)
    video_table_name = resolve_video_table(video_table)
    rows = conn.execute(
        f"""
        SELECT
            v.*,
            r.responses_json,
            r.progress_pct,
            r.is_complete,
            r.flag_for_review,
            r.updated_at
        FROM {video_table_name} v
        LEFT JOIN {table_name} r
            ON r.video_id = v.id
           AND r.researcher_id = ?
        ORDER BY v.source_row ASC
        """,
        (researcher_id,),
    ).fetchall()
    allowed_ids = set(video_ids) if video_ids else None
    queue_rows = []
    for row in rows:
        if allowed_ids is not None and row["id"] not in allowed_ids:
            continue
        responses = parse_responses(row["responses_json"])
        progress_pct = round(float(row["progress_pct"] or 0), 1)
        status = "not-started"
        if row["is_complete"]:
            status = "complete"
        elif progress_pct > 0:
            status = "in-progress"
        queue_rows.append(
            {
                "id": row["id"],
                "creator_handle": row["creator_handle"],
                "description": row["description"],
                "tiktok_url": row["tiktok_url"],
                "post_date": row["post_date"],
                "views": row["views"],
                "likes": row["likes"],
                "comments": row["comments"],
                "shares": row["shares"],
                "progress_pct": progress_pct,
                "status": status,
                "updated_at": row["updated_at"],
                "misinfo_score": responses.get("B3_Misinfo_Score"),
                "flag_for_review": bool(row["flag_for_review"]),
            }
        )
    return queue_rows


def build_researcher_stats(
    conn: sqlite3.Connection,
    slug: str,
    table: str = PRIMARY_RATING_TABLE,
    video_ids: list[str] | None = None,
    video_table: str = PRIMARY_VIDEO_TABLE,
) -> dict[str, Any]:
    researcher = fetch_researcher_by_slug(conn, slug)
    if researcher is None:
        return {}
    rows = fetch_queue_rows(conn, researcher["id"], table=table, video_ids=video_ids, video_table=video_table)
    total = len(rows)
    completed = sum(1 for row in rows if row["status"] == "complete")
    in_progress = sum(1 for row in rows if row["status"] == "in-progress")
    flagged = sum(1 for row in rows if row["flag_for_review"])
    progress_pct = round((sum(float(row["progress_pct"]) for row in rows) / total), 1) if total else 0
    latest = None
    for row in rows:
        if row["updated_at"] and (latest is None or row["updated_at"] > latest):
            latest = row["updated_at"]
    return {
        "total": total,
        "completed": completed,
        "in_progress": in_progress,
        "flagged": flagged,
        "progress_pct": progress_pct,
        "latest_activity": latest,
    }


def export_rows_for_researcher(
    conn: sqlite3.Connection,
    researcher_id: int,
    researcher_name: str,
    table: str = PRIMARY_RATING_TABLE,
    video_ids: list[str] | None = None,
    video_table: str = PRIMARY_VIDEO_TABLE,
) -> list[dict[str, Any]]:
    table_name = resolve_rating_table(table)
    video_table_name = resolve_video_table(video_table)
    rows = conn.execute(
        f"""
        SELECT
            v.*,
            r.responses_json,
            r.updated_at,
            r.progress_pct,
            r.is_complete
        FROM {video_table_name} v
        LEFT JOIN {table_name} r
            ON r.video_id = v.id
           AND r.researcher_id = ?
        ORDER BY v.source_row ASC
        """,
        (researcher_id,),
    ).fetchall()
    allowed_ids = set(video_ids) if video_ids else None
    export_rows = []
    for row in rows:
        if allowed_ids is not None and row["id"] not in allowed_ids:
            continue
        responses = parse_responses(row["responses_json"])
        export_row: dict[str, Any] = {
            "Coder_Name": researcher_name,
            "Last_Updated_UTC": row["updated_at"] or "",
            "Progress_Pct": row["progress_pct"] or 0,
            "Is_Complete": row["is_complete"] or 0,
            "Video_ID": row["id"],
            "TikTok_URL": row["tiktok_url"],
            "Views": row["views"] or "",
            "Likes": row["likes"] or "",
            "Comments": row["comments"] or "",
            "Shares": row["shares"] or "",
            "Post_Date": row["post_date"] or "",
            "Creator_Handle": row["creator_handle"] or "",
        }
        for field_id in FIELD_ORDER:
            value = responses.get(field_id, "")
            if isinstance(value, list):
                value = ", ".join(value)
            export_row[field_id] = value
        export_row["DISCERN_Total"] = compute_discern_total(responses) or ""
        export_rows.append(export_row)
    return export_rows


def csv_response(rows: list[dict[str, Any]], filename: str) -> Any:
    output = io.StringIO()
    columns = ["Coder_Name", "Last_Updated_UTC", "Progress_Pct", "Is_Complete"] + EXPORT_FIELD_ORDER
    writer = csv.DictWriter(output, fieldnames=columns)
    writer.writeheader()
    writer.writerows(rows)
    response = make_response(output.getvalue())
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def write_csv_snapshot(conn: sqlite3.Connection, destination: Path) -> None:
    rows: list[dict[str, Any]] = []
    for researcher in fetch_researchers(conn):
        rows.extend(export_rows_for_researcher(conn, researcher["id"], researcher["display_name"]))
    columns = ["Coder_Name", "Last_Updated_UTC", "Progress_Pct", "Is_Complete"] + EXPORT_FIELD_ORDER
    with destination.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def write_backup_snapshot(conn: sqlite3.Connection, reason: str = "manual") -> Path | None:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    latest = latest_backup_snapshot()
    if latest is not None and BACKUP_MIN_INTERVAL_SECONDS > 0:
        age_seconds = datetime.utcnow().timestamp() - latest.stat().st_mtime
        if age_seconds < BACKUP_MIN_INTERVAL_SECONDS:
            return None
    stamp = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    prefix = f"{stamp}_{reason}"
    sqlite_path = BACKUP_DIR / f"{prefix}.sqlite3"
    csv_path = BACKUP_DIR / f"{prefix}_all.csv"
    meta_path = BACKUP_DIR / f"{prefix}.json"
    snapshot_conn = sqlite3.connect(sqlite_path)
    try:
        conn.backup(snapshot_conn)
    finally:
        snapshot_conn.close()
    write_csv_snapshot(conn, csv_path)
    meta_path.write_text(
        json.dumps(
            {
                "created_at_utc": utcnow(),
                "reason": reason,
                "ratings_count": conn.execute("SELECT COUNT(*) AS count FROM ratings").fetchone()["count"],
                "calibration_ratings_count": conn.execute("SELECT COUNT(*) AS count FROM calibration_ratings").fetchone()["count"],
                "researchers_count": conn.execute("SELECT COUNT(*) AS count FROM researchers").fetchone()["count"],
                "videos_count": conn.execute("SELECT COUNT(*) AS count FROM videos").fetchone()["count"],
                "calibration_videos_count": conn.execute("SELECT COUNT(*) AS count FROM calibration_videos").fetchone()["count"],
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    prune_backups()
    return sqlite_path


def latest_backup_snapshot() -> Path | None:
    sqlite_backups = sorted(BACKUP_DIR.glob("*.sqlite3"), key=lambda path: path.stat().st_mtime, reverse=True)
    return sqlite_backups[0] if sqlite_backups else None


def prune_backups() -> None:
    sqlite_backups = sorted(BACKUP_DIR.glob("*.sqlite3"), key=lambda path: path.stat().st_mtime, reverse=True)
    if len(sqlite_backups) <= BACKUP_MAX_FILES:
        return
    for sqlite_path in sqlite_backups[BACKUP_MAX_FILES:]:
        prefix = sqlite_path.stem
        for related in BACKUP_DIR.glob(f"{prefix}*"):
            related.unlink(missing_ok=True)


def fetch_calibration_video_ids(conn: sqlite3.Connection) -> list[str]:
    rows = conn.execute(
        f"SELECT id FROM {CALIBRATION_VIDEO_TABLE} ORDER BY source_row ASC"
    ).fetchall()
    return [row["id"] for row in rows]


def fetch_videos_by_ids(
    conn: sqlite3.Connection,
    video_ids: list[str],
    video_table: str = PRIMARY_VIDEO_TABLE,
) -> list[sqlite3.Row]:
    if not video_ids:
        return []
    video_table_name = resolve_video_table(video_table)
    order_lookup = {video_id: index for index, video_id in enumerate(video_ids)}
    rows = conn.execute(f"SELECT * FROM {video_table_name} ORDER BY source_row ASC").fetchall()
    filtered_rows = [row for row in rows if row["id"] in order_lookup]
    filtered_rows.sort(key=lambda row: order_lookup[row["id"]])
    return filtered_rows


def save_rating_responses(
    conn: sqlite3.Connection,
    researcher_id: int,
    video_id: str,
    responses: dict[str, Any],
    table: str = PRIMARY_RATING_TABLE,
) -> dict[str, Any]:
    table_name = resolve_rating_table(table)
    progress_pct = compute_progress_pct(responses)
    is_complete = int(progress_pct == 100)
    flag_for_review = int(responses.get("Flag_For_Review") == "1")
    now = utcnow()
    existing = conn.execute(
        f"SELECT id, started_at, completed_at FROM {table_name} WHERE researcher_id = ? AND video_id = ?",
        (researcher_id, video_id),
    ).fetchone()
    if existing is None:
        started_at = now if responses else None
        completed_at = now if is_complete else None
        conn.execute(
            f"""
            INSERT INTO {table_name} (
                researcher_id, video_id, responses_json, progress_pct, is_complete,
                flag_for_review, started_at, updated_at, completed_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                researcher_id,
                video_id,
                json.dumps(responses),
                progress_pct,
                is_complete,
                flag_for_review,
                started_at,
                now,
                completed_at,
            ),
        )
    else:
        started_at = existing["started_at"] or (now if responses else None)
        completed_at = existing["completed_at"]
        if is_complete and not completed_at:
            completed_at = now
        if not is_complete:
            completed_at = None
        conn.execute(
            f"""
            UPDATE {table_name}
            SET responses_json = ?, progress_pct = ?, is_complete = ?,
                flag_for_review = ?, started_at = ?, updated_at = ?, completed_at = ?
            WHERE researcher_id = ? AND video_id = ?
            """,
            (
                json.dumps(responses),
                progress_pct,
                is_complete,
                flag_for_review,
                started_at,
                now,
                completed_at,
                researcher_id,
                video_id,
            ),
        )
    conn.commit()
    return {
        "ok": True,
        "progress_pct": progress_pct,
        "is_complete": bool(is_complete),
        "discern_total": compute_discern_total(responses),
        "missing_required": required_field_ids(responses),
        "updated_at": now,
    }


def build_calibration_comparison_rows(
    conn: sqlite3.Connection,
    video_ids: list[str],
    researchers: list[sqlite3.Row],
) -> list[dict[str, Any]]:
    ratings = fetch_all_ratings(conn, table=CALIBRATION_RATING_TABLE, video_ids=video_ids)
    rating_lookup = {(row["researcher_id"], row["video_id"]): row for row in ratings}
    comparison_rows: list[dict[str, Any]] = []
    for video in fetch_videos_by_ids(conn, video_ids, video_table=CALIBRATION_VIDEO_TABLE):
        coder_rows = []
        completed_count = 0
        misinfo_values: list[int] = []
        discern_totals: list[int] = []
        for researcher in researchers:
            rating = rating_lookup.get((researcher["id"], video["id"]))
            status = "not-started"
            misinfo_score = None
            discern_total = None
            updated_at = None
            if rating is not None:
                status = "complete" if rating["is_complete"] else "in-progress" if rating["progress_pct"] > 0 else "not-started"
                responses = rating["responses"]
                misinfo_score = responses.get("B3_Misinfo_Score")
                discern_total = compute_discern_total(responses)
                updated_at = rating["updated_at"]
                if rating["is_complete"]:
                    completed_count += 1
                    if misinfo_score in {"1", "2", "3", "4", "5"}:
                        misinfo_values.append(int(misinfo_score))
                    if discern_total is not None:
                        discern_totals.append(discern_total)
            coder_rows.append(
                {
                    "display_name": researcher["display_name"],
                    "status": status,
                    "misinfo_score": misinfo_score,
                    "discern_total": discern_total,
                    "updated_at": updated_at,
                }
            )
        misinfo_spread = (max(misinfo_values) - min(misinfo_values)) if len(misinfo_values) >= 2 else 0
        discern_spread = (max(discern_totals) - min(discern_totals)) if len(discern_totals) >= 2 else 0
        needs_discussion = misinfo_spread > 0 or discern_spread >= CALIBRATION_DISCERN_SPREAD_THRESHOLD
        comparison_rows.append(
            {
                "video": video,
                "coders": coder_rows,
                "completed_count": completed_count,
                "total_coders": len(researchers),
                "misinfo_spread": misinfo_spread,
                "discern_spread": discern_spread,
                "needs_discussion": needs_discussion,
            }
        )
    return comparison_rows


def hydrate_rating_row(row: sqlite3.Row) -> dict[str, Any]:
    responses = parse_responses(row["responses_json"])
    return {
        "id": row["id"],
        "researcher_id": row["researcher_id"],
        "video_id": row["video_id"],
        "responses": responses,
        "progress_pct": round(float(row["progress_pct"] or compute_progress_pct(responses)), 1),
        "is_complete": bool(row["is_complete"]),
        "flag_for_review": bool(row["flag_for_review"]),
        "updated_at": row["updated_at"],
        "started_at": row["started_at"],
        "completed_at": row["completed_at"],
    }


def parse_responses(payload: str | None) -> dict[str, Any]:
    if not payload:
        return {}
    try:
        parsed = json.loads(payload)
    except json.JSONDecodeError:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def sanitize_responses(payload: dict[str, Any]) -> dict[str, Any]:
    clean: dict[str, Any] = {}
    for field_id, field in FIELD_INDEX.items():
        raw_value = payload.get(field_id)
        if field["type"] == "textarea":
            text = " ".join(str(raw_value).split()) if raw_value else ""
            if text:
                clean[field_id] = text
            continue
        if field["type"] == "multi":
            values = raw_value if isinstance(raw_value, list) else []
            allowed = {option["value"] for option in field["options"]}
            selected = []
            for value in values:
                value_str = str(value).strip()
                if value_str in allowed and value_str not in selected:
                    selected.append(value_str)
            if selected:
                clean[field_id] = selected
            continue
        value = "" if raw_value is None else str(raw_value).strip()
        allowed = {option["value"] for option in field["options"]}
        if value in allowed:
            clean[field_id] = value
    return clean


def compute_discern_total(responses: dict[str, Any]) -> int | None:
    values = []
    for field_id in DISCERN_FIELDS:
        value = responses.get(field_id)
        if value in {"1", "2", "3", "4", "5"}:
            values.append(int(value))
    if not values:
        return None
    return sum(values)


def required_field_ids(responses: dict[str, Any]) -> list[str]:
    required = []
    for field_id, field in FIELD_INDEX.items():
        if field_is_required(field, responses):
            required.append(field_id)
    return [field_id for field_id in required if not field_answered(field_id, responses)]


def field_is_required(field: dict[str, Any], responses: dict[str, Any]) -> bool:
    if field.get("required"):
        return True
    requirement = field.get("required_when")
    if not requirement:
        return False
    controlling_value = responses.get(requirement["field"])
    return controlling_value in requirement["values"]


def field_answered(field_id: str, responses: dict[str, Any]) -> bool:
    field = FIELD_INDEX[field_id]
    value = responses.get(field_id)
    if field["type"] == "textarea":
        return bool(value)
    if field["type"] == "multi":
        return bool(value)
    return value in {option["value"] for option in field["options"]}


def compute_progress_pct(responses: dict[str, Any]) -> float:
    required_ids = [field_id for field_id, field in FIELD_INDEX.items() if field_is_required(field, responses)]
    if not required_ids:
        return 0.0
    answered = sum(1 for field_id in required_ids if field_answered(field_id, responses))
    return round((answered / len(required_ids)) * 100, 1)


def slugify(name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    return slug or "researcher"


def researcher_is_deletable(slug: str | None) -> bool:
    default_slugs = {item[0] for item in DEFAULT_RESEARCHERS}
    return bool(slug) and slug not in default_slugs


def unique_slug(conn: sqlite3.Connection, base_slug: str) -> str:
    slug = base_slug
    suffix = 2
    while conn.execute("SELECT 1 FROM researchers WHERE slug = ?", (slug,)).fetchone():
        slug = f"{base_slug}-{suffix}"
        suffix += 1
    return slug


def configure_sqlite_connection(conn: sqlite3.Connection) -> None:
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA synchronous = FULL")
    conn.execute("PRAGMA busy_timeout = 5000")


def sync_default_researchers(conn: sqlite3.Connection) -> None:
    now = utcnow()
    existing_count = conn.execute("SELECT COUNT(*) AS count FROM researchers").fetchone()["count"]
    if existing_count == 0:
        conn.executemany(
            """
            INSERT INTO researchers (slug, display_name, created_at, updated_at)
            VALUES (?, ?, ?, ?)
            """,
            [(slug, name, now, now) for slug, name in DEFAULT_RESEARCHERS],
        )
        conn.commit()
        return
    for slug, name in DEFAULT_RESEARCHERS:
        present = conn.execute("SELECT id FROM researchers WHERE slug = ?", (slug,)).fetchone()
        if present is None:
            conn.execute(
                """
                INSERT INTO researchers (slug, display_name, created_at, updated_at)
                VALUES (?, ?, ?, ?)
                """,
                (slug, name, now, now),
            )
    conn.commit()


def utcnow() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat()


def to_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    return int(value)


def extract_tiktok_numeric_id(url: str) -> str | None:
    match = re.search(r"/video/(\d+)", url or "")
    return match.group(1) if match else None


def build_embed_url(tiktok_numeric_id: str | None) -> str | None:
    if not tiktok_numeric_id:
        return None
    return f"https://www.tiktok.com/player/v1/{tiktok_numeric_id}?controls=1&description=0"


def pick_resume_video(videos: list[dict[str, Any]]) -> dict[str, Any] | None:
    in_progress = [video for video in videos if video["status"] == "in-progress"]
    if in_progress:
        return sorted(in_progress, key=lambda item: item["updated_at"] or "", reverse=True)[0]
    for video in videos:
        if video["status"] != "complete":
            return video
    return videos[0] if videos else None


def normalize_creator_handle(username: str | None) -> str:
    handle = str(username or "").strip()
    if not handle:
        return ""
    return handle if handle.startswith("@") else f"@{handle}"


def build_tiktok_url(username: str | None, numeric_id: str | None) -> str:
    clean_id = str(numeric_id or "").strip()
    clean_username = str(username or "").strip().lstrip("@")
    if clean_id and clean_username:
        return f"https://www.tiktok.com/@{clean_username}/video/{clean_id}"
    if clean_id:
        return f"https://www.tiktok.com/video/{clean_id}"
    return ""


def format_candidate_post_date(value: Any) -> str:
    try:
        timestamp = int(value)
    except (TypeError, ValueError):
        return ""
    return datetime.utcfromtimestamp(timestamp).strftime("%Y-%m-%d")


if __name__ == "__main__":
    app = create_app()
    runtime = current_runtime()
    port = int(os.environ.get("RESEARCH_DASHBOARD_PORT") or runtime.raw_config["study"].get("port", 5173))
    app.run(debug=True, port=port)
