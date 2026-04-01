from __future__ import annotations

import argparse
import json
import re
import shutil
import sqlite3
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .config import load_runtime
from .models import StudyRuntime


METADATA_COLUMNS = [
    "Video_ID",
    "TikTok_URL",
    "Views",
    "Likes",
    "Comments",
    "Shares",
    "Post_Date",
    "Creator_Handle",
]


@dataclass(frozen=True)
class ProjectConfig:
    workbook_paths: tuple[Path, ...]
    database_path: Path
    candidate_pool_paths: tuple[Path, ...]
    transcripts_dir: Path
    llm_scores_dir: Path
    archive_root: Path
    replacement_log_path: Path

    @property
    def primary_workbook_path(self) -> Path:
        for path in self.workbook_paths:
            if path.exists():
                return path
        return self.workbook_paths[0]


def project_config_from_runtime(runtime: StudyRuntime) -> ProjectConfig:
    return ProjectConfig(
        workbook_paths=(runtime.paths.workbook_path,),
        database_path=runtime.paths.database_path,
        candidate_pool_paths=(runtime.paths.metadata_json_path, runtime.paths.replacement_candidates_path),
        transcripts_dir=runtime.paths.transcripts_dir,
        llm_scores_dir=runtime.paths.llm_scores_dir,
        archive_root=runtime.paths.data_dir / "sample_replacement_archives",
        replacement_log_path=runtime.paths.replacement_log_path,
    )


def utc_timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def load_collected_videos(config: ProjectConfig) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    loaded_any = False
    for path in config.candidate_pool_paths:
        if not path.exists():
            continue
        loaded_any = True
        payload = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(payload, list):
            items.extend(payload)
    if not loaded_any:
        raise FileNotFoundError(
            "Missing candidate pool files: " + ", ".join(str(path) for path in config.candidate_pool_paths)
        )
    return items


def load_collected_lookup(config: ProjectConfig) -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    for item in load_collected_videos(config):
        item_id = str(item.get("id") or "").strip()
        if item_id:
            lookup[item_id] = item
    return lookup


def connect_db(database_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(database_path, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout = 30000")
    return conn


def extract_numeric_id_from_url(url: str | None) -> str:
    if not url:
        return ""
    match = re.search(r"/video/(\d+)", str(url))
    return match.group(1) if match else ""


def format_date_from_epoch(value: Any) -> str:
    if not value:
        return ""
    return datetime.fromtimestamp(int(value), tz=timezone.utc).strftime("%Y-%m-%d")


def normalize_creator_handle(username: str | None) -> str:
    if not username:
        return ""
    username = str(username).strip()
    if not username:
        return ""
    return username if username.startswith("@") else f"@{username}"


def build_candidate_record(item: dict[str, Any]) -> dict[str, Any]:
    numeric_id = str(item.get("id") or "").strip()
    username = str(item.get("username") or "").strip()
    return {
        "tiktok_numeric_id": numeric_id,
        "tiktok_url": f"https://www.tiktok.com/@{username}/video/{numeric_id}" if username and numeric_id else "",
        "views": int(item.get("view_count") or 0),
        "likes": int(item.get("like_count") or 0),
        "comments": int(item.get("comment_count") or 0),
        "shares": int(item.get("share_count") or 0),
        "post_date": format_date_from_epoch(item.get("create_time")),
        "creator_handle": normalize_creator_handle(username),
        "description": str(item.get("video_description") or ""),
        "username": username,
        "region_code": str(item.get("region_code") or ""),
    }


def get_workbook_headers(sheet: Any) -> list[str]:
    return [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]


def get_used_numeric_ids(config: ProjectConfig) -> set[str]:
    primary_workbook = config.primary_workbook_path
    if not primary_workbook.exists():
        return set()
    workbook = load_workbook(primary_workbook, data_only=True)
    try:
        sheet = workbook["Reviewer_A"]
        headers = get_workbook_headers(sheet)
        url_index = headers.index("TikTok_URL") + 1
        used_ids = set()
        for row_index in range(4, sheet.max_row + 1):
            numeric_id = extract_numeric_id_from_url(sheet.cell(row_index, url_index).value)
            if numeric_id:
                used_ids.add(numeric_id)
        return used_ids
    finally:
        workbook.close()


def list_reserve_candidates(config: ProjectConfig) -> list[dict[str, Any]]:
    used_ids = get_used_numeric_ids(config)
    items = []
    for item in load_collected_videos(config):
        numeric_id = str(item.get("id") or "").strip()
        if numeric_id and numeric_id not in used_ids:
            items.append(build_candidate_record(item))
    return sorted(items, key=lambda item: item["views"], reverse=True)


def inspect_slot(slot_id: str, config: ProjectConfig) -> dict[str, Any]:
    slot_id = slot_id.upper()
    workbook_path = config.primary_workbook_path
    if not workbook_path.exists():
        raise FileNotFoundError(f"Missing workbook: {workbook_path}")
    workbook = load_workbook(workbook_path, data_only=True)
    try:
        sheet = workbook["Reviewer_A"]
        headers = get_workbook_headers(sheet)
        row_index = None
        workbook_row = None
        for candidate_row in range(4, sheet.max_row + 1):
            if str(sheet.cell(candidate_row, 1).value or "").upper() == slot_id:
                row_index = candidate_row
                workbook_row = {
                    headers[column - 1]: sheet.cell(candidate_row, column).value
                    for column in range(1, min(sheet.max_column, 8) + 1)
                }
                break
    finally:
        workbook.close()
    if row_index is None or workbook_row is None:
        raise ValueError(f"Could not find slot {slot_id} in {workbook_path}")
    transcript_path = config.transcripts_dir / f"{slot_id}.txt"
    llm_score_path = config.llm_scores_dir / slot_id
    conn = connect_db(config.database_path)
    try:
        video_row = conn.execute(
            """
            SELECT id, tiktok_url, tiktok_numeric_id, views, likes, comments, shares,
                   post_date, creator_handle, length(coalesce(description, '')) AS description_length,
                   length(coalesce(transcript, '')) AS transcript_length
            FROM videos
            WHERE id = ?
            """,
            (slot_id,),
        ).fetchone()
        rating_rows = conn.execute(
            """
            SELECT rs.display_name, r.progress_pct, r.is_complete, r.updated_at
            FROM ratings r
            JOIN researchers rs ON rs.id = r.researcher_id
            WHERE r.video_id = ? AND (r.progress_pct > 0 OR r.responses_json != '{}')
            ORDER BY rs.display_name COLLATE NOCASE
            """,
            (slot_id,),
        ).fetchall()
    finally:
        conn.close()
    return {
        "slot_id": slot_id,
        "source_row": row_index,
        "workbook": workbook_row,
        "database": dict(video_row) if video_row else None,
        "ratings": [dict(row) for row in rating_rows],
        "transcript_exists": transcript_path.exists(),
        "llm_scores_exist": llm_score_path.exists(),
    }


def backup_database(database_path: Path, backup_path: Path) -> None:
    backup_path.parent.mkdir(parents=True, exist_ok=True)
    src = connect_db(database_path)
    dest = sqlite3.connect(backup_path)
    try:
        src.backup(dest)
    finally:
        dest.close()
        src.close()


def backup_workbook(workbook_path: Path, backup_path: Path) -> None:
    backup_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(workbook_path, backup_path)


def archive_if_exists(source: Path, destination: Path) -> str | None:
    if not source.exists():
        return None
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(str(source), str(destination))
    return str(destination)


def clear_coder_cells(sheet: Any, row_index: int, headers: list[str]) -> None:
    if "Pre_Watch_Prediction" in headers:
        start_column = headers.index("Pre_Watch_Prediction") + 1
    elif "Creator_Handle" in headers:
        start_column = headers.index("Creator_Handle") + 2
    else:
        start_column = min(len(METADATA_COLUMNS) + 1, sheet.max_column + 1)
    for column in range(start_column, sheet.max_column + 1):
        sheet.cell(row_index, column).value = None


def update_workbook_slot(workbook_path: Path, slot_id: str, replacement: dict[str, Any]) -> dict[str, Any]:
    workbook = load_workbook(workbook_path)
    try:
        updated_sheets = []
        for sheet_name in ("Reviewer_A", "Reviewer_B"):
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            headers = get_workbook_headers(sheet)
            row_index = None
            for candidate_row in range(4, sheet.max_row + 1):
                if str(sheet.cell(candidate_row, 1).value or "").upper() == slot_id:
                    row_index = candidate_row
                    break
            if row_index is None:
                continue
            sheet.cell(row_index, headers.index("TikTok_URL") + 1).hyperlink = replacement["tiktok_url"]
            for field_name, value in (
                ("Video_ID", slot_id),
                ("TikTok_URL", replacement["tiktok_url"]),
                ("Views", replacement["views"]),
                ("Likes", replacement["likes"]),
                ("Comments", replacement["comments"]),
                ("Shares", replacement["shares"]),
                ("Post_Date", replacement["post_date"]),
                ("Creator_Handle", replacement["creator_handle"]),
            ):
                sheet.cell(row_index, headers.index(field_name) + 1).value = value
            clear_coder_cells(sheet, row_index, headers)
            updated_sheets.append({"sheet_name": sheet_name, "row_index": row_index})
        if not updated_sheets:
            raise ValueError(f"Could not find slot {slot_id} in workbook {workbook_path}")
        workbook.save(workbook_path)
        return {"workbook_path": str(workbook_path), "updated_sheets": updated_sheets}
    finally:
        workbook.close()


def update_database_slot(database_path: Path, slot_id: str, replacement: dict[str, Any]) -> dict[str, Any]:
    conn = connect_db(database_path)
    try:
        existing = conn.execute(
            """
            SELECT id, tiktok_url, tiktok_numeric_id, views, likes, comments, shares,
                   post_date, creator_handle, description
            FROM videos
            WHERE id = ?
            """,
            (slot_id,),
        ).fetchone()
        if existing is None:
            raise ValueError(f"Slot {slot_id} does not exist in database {database_path}")
        tables_to_clear = ["ratings"]
        available_tables = {
            row["name"] for row in conn.execute("SELECT name FROM sqlite_master WHERE type = 'table'").fetchall()
        }
        if "calibration_ratings" in available_tables:
            tables_to_clear.append("calibration_ratings")
        cleared_counts = {
            table_name: conn.execute(
                f"SELECT COUNT(*) AS count FROM {table_name} WHERE video_id = ?",
                (slot_id,),
            ).fetchone()["count"]
            for table_name in tables_to_clear
        }
        conn.execute("BEGIN IMMEDIATE")
        conn.execute(
            """
            UPDATE videos
            SET tiktok_url = ?,
                tiktok_numeric_id = ?,
                views = ?,
                likes = ?,
                comments = ?,
                shares = ?,
                post_date = ?,
                creator_handle = ?,
                description = ?,
                transcript = ''
            WHERE id = ?
            """,
            (
                replacement["tiktok_url"],
                replacement["tiktok_numeric_id"],
                replacement["views"],
                replacement["likes"],
                replacement["comments"],
                replacement["shares"],
                replacement["post_date"],
                replacement["creator_handle"],
                replacement["description"],
                slot_id,
            ),
        )
        for table_name in tables_to_clear:
            conn.execute(f"DELETE FROM {table_name} WHERE video_id = ?", (slot_id,))
        conn.commit()
        return {
            "cleared_rating_count": sum(cleared_counts.values()),
            "cleared_counts": cleared_counts,
            "previous_video": dict(existing),
        }
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def append_replacement_log(log_path: Path, entry: dict[str, Any]) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(entry, sort_keys=True) + "\n")


def replace_slot(
    slot_id: str,
    candidate_numeric_id: str,
    config: ProjectConfig,
    *,
    allow_existing_candidate: bool = False,
) -> dict[str, Any]:
    slot_id = slot_id.upper()
    candidate_numeric_id = str(candidate_numeric_id).strip()
    collected_lookup = load_collected_lookup(config)
    if candidate_numeric_id not in collected_lookup:
        raise ValueError(
            f"Candidate {candidate_numeric_id} was not found in "
            + ", ".join(str(path) for path in config.candidate_pool_paths)
        )
    used_ids = get_used_numeric_ids(config)
    current_state = inspect_slot(slot_id, config)
    current_numeric_id = extract_numeric_id_from_url(current_state["workbook"]["TikTok_URL"])
    if candidate_numeric_id != current_numeric_id and candidate_numeric_id in used_ids and not allow_existing_candidate:
        raise ValueError(f"Candidate {candidate_numeric_id} is already used in the workbook sample")
    replacement = build_candidate_record(collected_lookup[candidate_numeric_id])
    timestamp = utc_timestamp()
    backup_root = config.archive_root / "backups" / timestamp
    workbook_backups = []
    for workbook_path in config.workbook_paths:
        if workbook_path.exists():
            backup_path = backup_root / "workbooks" / workbook_path.name
            backup_workbook(workbook_path, backup_path)
            workbook_backups.append(str(backup_path))
    database_backup_path = backup_root / "database" / config.database_path.name
    backup_database(config.database_path, database_backup_path)
    workbook_updates = []
    for workbook_path in config.workbook_paths:
        if workbook_path.exists():
            workbook_updates.append(update_workbook_slot(workbook_path, slot_id, replacement))
    database_update = update_database_slot(config.database_path, slot_id, replacement)
    transcript_archive = archive_if_exists(
        config.transcripts_dir / f"{slot_id}.txt",
        config.archive_root / "transcripts" / f"{slot_id}_{timestamp}.txt",
    )
    llm_scores_archive = archive_if_exists(
        config.llm_scores_dir / slot_id,
        config.archive_root / "llm_scores" / f"{slot_id}_{timestamp}",
    )
    log_entry = {
        "timestamp_utc": timestamp,
        "slot_id": slot_id,
        "candidate_numeric_id": candidate_numeric_id,
        "replacement": replacement,
        "previous_state": current_state,
        "database_backup": str(database_backup_path),
        "workbook_backups": workbook_backups,
        "transcript_archive": transcript_archive,
        "llm_scores_archive": llm_scores_archive,
        "cleared_rating_count": database_update["cleared_rating_count"],
    }
    append_replacement_log(config.replacement_log_path, log_entry)
    return {
        "slot_id": slot_id,
        "candidate_numeric_id": candidate_numeric_id,
        "replacement": replacement,
        "current_state": current_state,
        "database_update": database_update,
        "workbook_updates": workbook_updates,
        "database_backup": str(database_backup_path),
        "workbook_backups": workbook_backups,
        "transcript_archive": transcript_archive,
        "llm_scores_archive": llm_scores_archive,
        "log_path": str(config.replacement_log_path),
    }


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Inspect and replace video slots in a study workspace.")
    parser.add_argument("--workspace", required=True, help="Workspace root containing study_config.yaml")
    subparsers = parser.add_subparsers(dest="command", required=True)
    inspect_parser = subparsers.add_parser("inspect-slot", help="Show the current state of a workbook slot")
    inspect_parser.add_argument("slot_id", help="Slot id like VID002")
    reserve_parser = subparsers.add_parser("list-reserve", help="List unused collected videos not in the workbook")
    reserve_parser.add_argument("--limit", type=int, default=20, help="Number of candidates to show")
    replace_parser = subparsers.add_parser("replace-slot", help="Replace a workbook slot with a reserve candidate")
    replace_parser.add_argument("slot_id", help="Slot id like VID002")
    replace_parser.add_argument("candidate_numeric_id", help="TikTok numeric id from workspace metadata")
    replace_parser.add_argument("--yes", action="store_true", help="Apply the replacement instead of previewing it")
    replace_parser.add_argument(
        "--allow-existing-candidate",
        action="store_true",
        help="Allow replacing with a candidate already used elsewhere in the workbook.",
    )
    return parser


def print_json(data: dict[str, Any] | list[dict[str, Any]]) -> None:
    print(json.dumps(data, indent=2, ensure_ascii=False))


def handle_inspect_slot(args: argparse.Namespace, config: ProjectConfig) -> int:
    print_json(inspect_slot(args.slot_id, config))
    return 0


def handle_list_reserve(args: argparse.Namespace, config: ProjectConfig) -> int:
    candidates = list_reserve_candidates(config)[: args.limit]
    print_json(candidates)
    return 0


def handle_replace_slot(args: argparse.Namespace, config: ProjectConfig) -> int:
    current_state = inspect_slot(args.slot_id, config)
    collected_lookup = load_collected_lookup(config)
    candidate_numeric_id = str(args.candidate_numeric_id).strip()
    if candidate_numeric_id not in collected_lookup:
        raise ValueError(
            f"Candidate {candidate_numeric_id} was not found in "
            + ", ".join(str(path) for path in config.candidate_pool_paths)
        )
    replacement = build_candidate_record(collected_lookup[candidate_numeric_id])
    preview = {
        "slot_id": current_state["slot_id"],
        "current_url": current_state["workbook"]["TikTok_URL"],
        "current_creator": current_state["workbook"]["Creator_Handle"],
        "current_numeric_id": extract_numeric_id_from_url(current_state["workbook"]["TikTok_URL"]),
        "candidate_numeric_id": candidate_numeric_id,
        "candidate_url": replacement["tiktok_url"],
        "candidate_creator": replacement["creator_handle"],
        "candidate_views": replacement["views"],
        "ratings_that_would_be_cleared": current_state["ratings"],
        "transcript_would_be_archived": current_state["transcript_exists"],
        "llm_scores_would_be_archived": current_state["llm_scores_exist"],
    }
    if not args.yes:
        print_json(preview)
        print("\nDry run only. Re-run with --yes to apply this replacement.", file=sys.stderr)
        return 2
    result = replace_slot(
        args.slot_id,
        candidate_numeric_id,
        config,
        allow_existing_candidate=args.allow_existing_candidate,
    )
    print_json(result)
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    runtime = load_runtime(Path(args.workspace))
    config = project_config_from_runtime(runtime)
    try:
        if args.command == "inspect-slot":
            return handle_inspect_slot(args, config)
        if args.command == "list-reserve":
            return handle_list_reserve(args, config)
        if args.command == "replace-slot":
            return handle_replace_slot(args, config)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1
    parser.error(f"Unsupported command: {args.command}")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
