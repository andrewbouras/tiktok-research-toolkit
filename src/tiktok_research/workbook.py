from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook

from .form_schema import DISCERN_FIELDS, EXPORT_METADATA_FIELDS, FIELD_INDEX, FIELD_ORDER, MISINFO_SCORE_GUIDE


def creator_handle(username: str | None) -> str:
    if not username:
        return ""
    return username if username.startswith("@") else f"@{username}"


def video_url(username: str | None, numeric_id: str | None) -> str:
    clean_username = str(username or "").strip().lstrip("@")
    clean_numeric_id = str(numeric_id or "").strip()
    if clean_username and clean_numeric_id:
        return f"https://www.tiktok.com/@{clean_username}/video/{clean_numeric_id}"
    if clean_numeric_id:
        return f"https://www.tiktok.com/video/{clean_numeric_id}"
    return ""


def write_coding_sheet(sheet: Any, videos: list[dict[str, Any]]) -> None:
    headers = EXPORT_METADATA_FIELDS + FIELD_ORDER + ["DISCERN_Total"]
    for index, header in enumerate(headers, start=1):
        sheet.cell(1, index).value = header
        if header in FIELD_INDEX:
            field = FIELD_INDEX[header]
            sheet.cell(2, index).value = field.get("help_text") or field.get("description") or ""
            sheet.cell(3, index).value = field.get("entry_hint") or ""

    for row_index, item in enumerate(videos, start=4):
        values = {
            "Video_ID": item["video_id"],
            "TikTok_URL": video_url(item.get("username"), item.get("id")),
            "Views": item.get("view_count", 0),
            "Likes": item.get("like_count", 0),
            "Comments": item.get("comment_count", 0),
            "Shares": item.get("share_count", 0),
            "Post_Date": item.get("post_date", ""),
            "Creator_Handle": creator_handle(item.get("username")),
        }
        for column, header in enumerate(headers, start=1):
            value = values.get(header)
            if value is not None:
                sheet.cell(row_index, column).value = value
                if header == "TikTok_URL" and value:
                    sheet.cell(row_index, column).hyperlink = value


def write_discern_reference(sheet: Any) -> None:
    sheet.cell(1, 1).value = "Question"
    sheet.cell(1, 2).value = "Field_ID"
    sheet.cell(1, 3).value = "Reference note"
    for index, field_id in enumerate(DISCERN_FIELDS, start=4):
        sheet.cell(index, 1).value = f"Q{index - 3}"
        sheet.cell(index, 2).value = field_id
        sheet.cell(index, 3).value = FIELD_INDEX[field_id].get("reference_note") or FIELD_INDEX[field_id].get("help_text") or ""


def write_clinical_reference(sheet: Any) -> None:
    sheet.cell(1, 1).value = "Score"
    sheet.cell(1, 2).value = "Label"
    sheet.cell(1, 3).value = "Definition"
    sheet.cell(1, 5).value = "Example"
    for row_index, item in enumerate(MISINFO_SCORE_GUIDE, start=6):
        sheet.cell(row_index, 1).value = item["score"]
        sheet.cell(row_index, 2).value = item["label"]
        sheet.cell(row_index, 3).value = item["definition"]
        sheet.cell(row_index, 5).value = item["example"]


def create_template_workbook(path: Path, videos: list[dict[str, Any]] | None = None) -> None:
    workbook = Workbook()
    reviewer_a = workbook.active
    reviewer_a.title = "Reviewer_A"
    reviewer_b = workbook.create_sheet("Reviewer_B")
    discern_reference = workbook.create_sheet("DISCERN_Reference")
    clinical_reference = workbook.create_sheet("CLINICAL_REFERENCE")

    write_coding_sheet(reviewer_a, videos or [])
    write_coding_sheet(reviewer_b, videos or [])
    write_discern_reference(discern_reference)
    write_clinical_reference(clinical_reference)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()

